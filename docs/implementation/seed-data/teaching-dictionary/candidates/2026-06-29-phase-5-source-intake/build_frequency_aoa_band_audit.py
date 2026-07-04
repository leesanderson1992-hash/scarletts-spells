#!/usr/bin/env python3
"""Build audit-only British frequency and AoA banding recommendations.

This script generates candidate review artifacts only. It does not update
canonical_words.csv, import data, touch Supabase, or change runtime behavior.

Optional AoA source files can be supplied later with:
- KUPERMAN_AOA_PATH=/path/to/file.csv|tsv|xlsx
- BRYSBAERT_BIEMILLER_AOA_PATH=/path/to/file.csv|tsv|xlsx
"""

from __future__ import annotations

import csv
import json
import os
import re
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Any
from urllib.request import urlopen


BASE = Path(__file__).resolve().parent
CSV_DIR = BASE / "csv"
WORDS_PATH = CSV_DIR / "canonical_words.csv"
SOURCES_PATH = CSV_DIR / "teaching_content_sources.csv"
BUILD_SUMMARY_PATH = BASE / "build_summary.json"

BNC_SOURCE_CACHE = BASE / "british_frequency_bnc_1_2_all_freq_source.txt"
BRITISH_FREQUENCY_INTAKE_PATH = BASE / "british_frequency_intake_all_canonical_words.csv"
AOA_INTAKE_PATH = BASE / "aoa_intake_all_canonical_words.csv"
RECOMMENDATIONS_PATH = BASE / "canonical_words_frequency_aoa_band_recommendations.csv"
SUMMARY_PATH = BASE / "frequency_aoa_band_audit_summary.json"

BNC_SOURCE_KEY = "bnc_frequency_lists_ucrel_2026_07_01"
BNC_SOURCE_NAME = "BNC Frequency Lists"
BNC_SOURCE_URL = "http://ucrel.lancs.ac.uk/bncfreq/lists/1_2_all_freq.txt"
BNC_SOURCE_PAGE_URL = "https://ucrel.lancs.ac.uk/bncfreq/"
BNC_SOURCE_LICENCE = "Creative Commons Attribution-ShareAlike 2.0 UK; legal review required before product import"
BNC_SOURCE_NOTE = (
    "British National Corpus frequency list used for audit-only frequency band "
    "recommendations. Active canonical_words.csv frequency_band values are not "
    "updated by this pass."
)

KUPERMAN_SOURCE_KEY = "kuperman_stadthagen_brysbaert_aoa_2012_reference_2026_07_01"
KUPERMAN_SOURCE_NAME = "Kuperman Stadthagen-Gonzalez Brysbaert AoA ratings"
KUPERMAN_SOURCE_URL = "https://link.springer.com/article/10.3758/s13428-012-0210-4"
KUPERMAN_SOURCE_LICENCE = "Article/supplementary data terms require legal review before product import"
KUPERMAN_SOURCE_NOTE = (
    "Reference source for adult-rated age-of-acquisition evidence. Source data "
    "file was not bundled by this pass unless KUPERMAN_AOA_PATH is supplied."
)

BRYSBAERT_BIEMILLER_SOURCE_KEY = "brysbaert_biemiller_aoa_2017_reference_2026_07_01"
BRYSBAERT_BIEMILLER_SOURCE_NAME = "Brysbaert Biemiller AoA norms"
BRYSBAERT_BIEMILLER_SOURCE_URL = "https://link.springer.com/article/10.3758/s13428-016-0811-4"
BRYSBAERT_BIEMILLER_SOURCE_LICENCE = (
    "Reported non-commercial research-use terms; legal review required before product import"
)
BRYSBAERT_BIEMILLER_SOURCE_NOTE = (
    "Reference source for educational age-of-acquisition/word-meaning evidence. "
    "Source data file was not bundled by this pass unless "
    "BRYSBAERT_BIEMILLER_AOA_PATH is supplied."
)

FREQUENCY_THRESHOLDS = {
    "high": 100.0,
    "medium": 10.0,
}


@dataclass(frozen=True)
class FrequencyEvidence:
    match_status: str
    matched_forms: tuple[str, ...]
    raw_frequency: float | None
    rank: int | None
    proposed_band: str
    review_notes: str


@dataclass(frozen=True)
class AoAEvidence:
    source_data_status: str
    kuperman_match_status: str
    kuperman_matched_form: str
    kuperman_raw_aoa: str
    brysbaert_biemiller_match_status: str
    brysbaert_biemiller_matched_form: str
    brysbaert_biemiller_raw_aoa: str
    proposed_age_band: str
    review_notes: str


def read_csv(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    with path.open(newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        return list(reader.fieldnames or []), list(reader)


def write_csv(path: Path, fieldnames: list[str], rows: list[dict[str, str]]) -> None:
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def normalize_lookup(value: str) -> str:
    return value.strip().lower().replace("’", "'")


def split_components(value: str) -> list[str]:
    pieces = re.split(r"[/\s-]+", normalize_lookup(value))
    return [piece for piece in pieces if piece]


def ensure_bnc_source() -> None:
    if BNC_SOURCE_CACHE.exists() and BNC_SOURCE_CACHE.stat().st_size > 0:
        return
    data = urlopen(BNC_SOURCE_URL, timeout=30).read().decode("utf-8", "replace")
    BNC_SOURCE_CACHE.write_text(data, encoding="utf-8")


def load_bnc_frequency() -> dict[str, dict[str, Any]]:
    ensure_bnc_source()
    totals: dict[str, float] = defaultdict(float)
    pos_tags: dict[str, set[str]] = defaultdict(set)
    with BNC_SOURCE_CACHE.open(encoding="utf-8") as f:
        reader = csv.DictReader(f, delimiter="\t")
        for row in reader:
            word = normalize_lookup(row.get("Word", "")).removesuffix("*")
            if not word:
                continue
            try:
                frequency = float(row.get("Freq", "") or 0)
            except ValueError:
                frequency = 0.0
            totals[word] += frequency
            pos = row.get("PoS", "").strip()
            if pos:
                pos_tags[word].add(pos)

    ranked_words = sorted(totals.items(), key=lambda item: (-item[1], item[0]))
    return {
        word: {
            "frequency": frequency,
            "rank": index + 1,
            "pos_tags": "|".join(sorted(pos_tags[word])),
        }
        for index, (word, frequency) in enumerate(ranked_words)
    }


def frequency_band(frequency: float | None) -> str:
    if frequency is None:
        return ""
    if frequency >= FREQUENCY_THRESHOLDS["high"]:
        return "high"
    if frequency >= FREQUENCY_THRESHOLDS["medium"]:
        return "medium"
    return "low"


def select_frequency_evidence(word: str, bnc: dict[str, dict[str, Any]]) -> FrequencyEvidence:
    lookup = normalize_lookup(word)
    if lookup in bnc:
        entry = bnc[lookup]
        frequency = float(entry["frequency"])
        return FrequencyEvidence(
            match_status="exact",
            matched_forms=(lookup,),
            raw_frequency=frequency,
            rank=int(entry["rank"]),
            proposed_band=frequency_band(frequency),
            review_notes="Exact BNC word-form match.",
        )

    components = split_components(word)
    if len(components) > 1 and all(component in bnc for component in components):
        total = sum(float(bnc[component]["frequency"]) for component in components)
        best_rank = min(int(bnc[component]["rank"]) for component in components)
        return FrequencyEvidence(
            match_status="component",
            matched_forms=tuple(components),
            raw_frequency=total,
            rank=best_rank,
            proposed_band=frequency_band(total),
            review_notes=(
                "No exact BNC row; component frequencies summed for audit only. "
                "Review before using for active frequency truth."
            ),
        )

    return FrequencyEvidence(
        match_status="unmatched",
        matched_forms=(),
        raw_frequency=None,
        rank=None,
        proposed_band="",
        review_notes="No exact or component BNC frequency match.",
    )


def find_column(headers: list[str], candidates: list[str]) -> str | None:
    lowered = {header.lower().strip(): header for header in headers}
    for candidate in candidates:
        if candidate.lower() in lowered:
            return lowered[candidate.lower()]
    for header in headers:
        normal = re.sub(r"[^a-z0-9]", "", header.lower())
        for candidate in candidates:
            if re.sub(r"[^a-z0-9]", "", candidate.lower()) == normal:
                return header
    return None


def load_optional_aoa(path: Path | None) -> dict[str, dict[str, str]]:
    if not path or not path.exists():
        return {}

    rows: list[dict[str, Any]]
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        try:
            from openpyxl import load_workbook
        except ModuleNotFoundError:
            return {}
        workbook = load_workbook(path, read_only=True, data_only=True)
        worksheet = workbook[workbook.sheetnames[0]]
        values = list(worksheet.iter_rows(values_only=True))
        if not values:
            return {}
        headers = [str(value).strip() if value is not None else "" for value in values[0]]
        rows = [
            {headers[i]: value for i, value in enumerate(row) if i < len(headers)}
            for row in values[1:]
        ]
    else:
        delimiter = "\t" if path.suffix.lower() in {".tsv", ".txt"} else ","
        with path.open(newline="", encoding="utf-8-sig") as f:
            rows = list(csv.DictReader(f, delimiter=delimiter))
            headers = list(rows[0].keys()) if rows else []

    if not rows:
        return {}
    headers = list(rows[0].keys())
    word_col = find_column(headers, ["word", "Word", "item", "Item", "lemma", "Lemma"])
    aoa_col = find_column(
        headers,
        [
            "Rating.Mean",
            "AoA",
            "aoa",
            "mean_aoa",
            "mean AoA",
            "aoa_years",
            "AoA_Kup",
        ],
    )
    if not word_col or not aoa_col:
        return {}

    data: dict[str, dict[str, str]] = {}
    for row in rows:
        word = normalize_lookup(str(row.get(word_col, "") or ""))
        value = str(row.get(aoa_col, "") or "").strip()
        if word and value and word not in data:
            data[word] = {"matched_form": word, "raw_aoa": value}
    return data


def age_band_from_aoa(raw_values: list[str]) -> str:
    numeric = []
    for value in raw_values:
        try:
            numeric.append(float(value))
        except ValueError:
            continue
    if not numeric:
        return ""
    aoa = min(numeric)
    if aoa <= 6:
        return "early_primary"
    if aoa <= 9:
        return "middle_primary"
    if aoa <= 12:
        return "upper_primary"
    return "later_review"


def select_aoa_evidence(
    word: str,
    kuperman: dict[str, dict[str, str]],
    brysbaert_biemiller: dict[str, dict[str, str]],
    kuperman_source_status: str,
    brysbaert_source_status: str,
) -> AoAEvidence:
    lookup = normalize_lookup(word)
    kup = kuperman.get(lookup)
    bb = brysbaert_biemiller.get(lookup)
    statuses = []
    if kuperman_source_status != "loaded":
        statuses.append(f"kuperman:{kuperman_source_status}")
    if brysbaert_source_status != "loaded":
        statuses.append(f"brysbaert_biemiller:{brysbaert_source_status}")
    source_status = ";".join(statuses) if statuses else "loaded"

    raw_values = []
    if kup:
        raw_values.append(kup["raw_aoa"])
    if bb:
        raw_values.append(bb["raw_aoa"])

    if not kup and not bb:
        return AoAEvidence(
            source_data_status=source_status,
            kuperman_match_status="unavailable" if kuperman_source_status != "loaded" else "unmatched",
            kuperman_matched_form="",
            kuperman_raw_aoa="",
            brysbaert_biemiller_match_status=(
                "unavailable" if brysbaert_source_status != "loaded" else "unmatched"
            ),
            brysbaert_biemiller_matched_form="",
            brysbaert_biemiller_raw_aoa="",
            proposed_age_band="",
            review_notes="AoA source data not available or no exact match; keep active age_band unchanged.",
        )

    return AoAEvidence(
        source_data_status=source_status,
        kuperman_match_status="exact" if kup else "unmatched",
        kuperman_matched_form=kup["matched_form"] if kup else "",
        kuperman_raw_aoa=kup["raw_aoa"] if kup else "",
        brysbaert_biemiller_match_status="exact" if bb else "unmatched",
        brysbaert_biemiller_matched_form=bb["matched_form"] if bb else "",
        brysbaert_biemiller_raw_aoa=bb["raw_aoa"] if bb else "",
        proposed_age_band=age_band_from_aoa(raw_values),
        review_notes="Exact AoA evidence found; review before updating active age_band.",
    )


def update_source_register() -> None:
    fieldnames, rows = read_csv(SOURCES_PATH)
    replacements = {
        BNC_SOURCE_KEY: {
            "source_key": BNC_SOURCE_KEY,
            "source_category": "open_licensed",
            "source_name": BNC_SOURCE_NAME,
            "source_url": BNC_SOURCE_PAGE_URL,
            "source_licence": BNC_SOURCE_LICENCE,
            "source_use_note": BNC_SOURCE_NOTE,
            "importability_status": "requires_legal_review",
            "legal_review_status": "required",
        },
        KUPERMAN_SOURCE_KEY: {
            "source_key": KUPERMAN_SOURCE_KEY,
            "source_category": "reference_only",
            "source_name": KUPERMAN_SOURCE_NAME,
            "source_url": KUPERMAN_SOURCE_URL,
            "source_licence": KUPERMAN_SOURCE_LICENCE,
            "source_use_note": KUPERMAN_SOURCE_NOTE,
            "importability_status": "requires_legal_review",
            "legal_review_status": "required",
        },
        BRYSBAERT_BIEMILLER_SOURCE_KEY: {
            "source_key": BRYSBAERT_BIEMILLER_SOURCE_KEY,
            "source_category": "reference_only",
            "source_name": BRYSBAERT_BIEMILLER_SOURCE_NAME,
            "source_url": BRYSBAERT_BIEMILLER_SOURCE_URL,
            "source_licence": BRYSBAERT_BIEMILLER_SOURCE_LICENCE,
            "source_use_note": BRYSBAERT_BIEMILLER_SOURCE_NOTE,
            "importability_status": "requires_legal_review",
            "legal_review_status": "required",
        },
    }
    rows = [row for row in rows if row.get("source_key") not in replacements]
    rows.extend(replacements.values())
    write_csv(SOURCES_PATH, fieldnames, rows)


def update_build_summary(summary: dict[str, Any]) -> None:
    if not BUILD_SUMMARY_PATH.exists():
        return
    data = json.loads(BUILD_SUMMARY_PATH.read_text(encoding="utf-8"))
    data.update(
        {
            "frequency_aoa_banding_audit_created": True,
            "frequency_aoa_banding_audit_rows": summary["rows"],
            "british_frequency_intake_rows": summary["rows"],
            "british_frequency_match_counts": summary["frequency_match_counts"],
            "aoa_intake_rows": summary["rows"],
            "aoa_source_data_status": summary["aoa_source_data_status"],
            "canonical_words_frequency_aoa_recommendation_rows": summary["rows"],
            "canonical_words_frequency_band_values_unchanged": True,
            "canonical_words_age_band_values_unchanged": True,
            "sources": summary["source_rows_after_update"],
        }
    )
    BUILD_SUMMARY_PATH.write_text(
        json.dumps(data, indent=2, sort_keys=True) + "\n", encoding="utf-8"
    )


def main() -> None:
    _word_fields, word_rows = read_csv(WORDS_PATH)
    if len(word_rows) != 874 or len({row["word_key"] for row in word_rows}) != 874:
        raise SystemExit("Expected 874 unique canonical word rows")

    bnc = load_bnc_frequency()
    kuperman_path = Path(os.environ["KUPERMAN_AOA_PATH"]) if os.environ.get("KUPERMAN_AOA_PATH") else None
    bb_path = (
        Path(os.environ["BRYSBAERT_BIEMILLER_AOA_PATH"])
        if os.environ.get("BRYSBAERT_BIEMILLER_AOA_PATH")
        else None
    )
    kuperman = load_optional_aoa(kuperman_path)
    bb = load_optional_aoa(bb_path)
    kuperman_status = "loaded" if kuperman else "source_file_not_supplied"
    bb_status = "loaded" if bb else "source_file_not_supplied"

    frequency_rows: list[dict[str, str]] = []
    aoa_rows: list[dict[str, str]] = []
    recommendation_rows: list[dict[str, str]] = []
    frequency_counts: Counter[str] = Counter()
    proposed_frequency_counts: Counter[str] = Counter()
    proposed_age_counts: Counter[str] = Counter()

    for word_row in word_rows:
        frequency = select_frequency_evidence(word_row["normalised_word"], bnc)
        aoa = select_aoa_evidence(
            word_row["normalised_word"],
            kuperman,
            bb,
            kuperman_status,
            bb_status,
        )
        frequency_counts[frequency.match_status] += 1
        proposed_frequency_counts[frequency.proposed_band or "blank"] += 1
        proposed_age_counts[aoa.proposed_age_band or "blank"] += 1

        frequency_rows.append(
            {
                "word_key": word_row["word_key"],
                "normalised_word": word_row["normalised_word"],
                "display_word": word_row["display_word"],
                "dialect_code": word_row["dialect_code"],
                "source_category": "open_licensed",
                "source_name": BNC_SOURCE_NAME,
                "source_url": BNC_SOURCE_PAGE_URL,
                "source_licence": BNC_SOURCE_LICENCE,
                "source_use_note": BNC_SOURCE_NOTE,
                "match_status": frequency.match_status,
                "matched_form": " | ".join(frequency.matched_forms),
                "raw_frequency_per_million": ""
                if frequency.raw_frequency is None
                else f"{frequency.raw_frequency:g}",
                "frequency_rank": "" if frequency.rank is None else str(frequency.rank),
                "current_frequency_band": word_row["frequency_band"],
                "proposed_frequency_band": frequency.proposed_band,
                "review_status": "in_review",
                "review_notes": frequency.review_notes,
            }
        )

        aoa_rows.append(
            {
                "word_key": word_row["word_key"],
                "normalised_word": word_row["normalised_word"],
                "display_word": word_row["display_word"],
                "dialect_code": word_row["dialect_code"],
                "source_data_status": aoa.source_data_status,
                "kuperman_source_name": KUPERMAN_SOURCE_NAME,
                "kuperman_source_url": KUPERMAN_SOURCE_URL,
                "kuperman_match_status": aoa.kuperman_match_status,
                "kuperman_matched_form": aoa.kuperman_matched_form,
                "kuperman_raw_aoa": aoa.kuperman_raw_aoa,
                "brysbaert_biemiller_source_name": BRYSBAERT_BIEMILLER_SOURCE_NAME,
                "brysbaert_biemiller_source_url": BRYSBAERT_BIEMILLER_SOURCE_URL,
                "brysbaert_biemiller_match_status": aoa.brysbaert_biemiller_match_status,
                "brysbaert_biemiller_matched_form": aoa.brysbaert_biemiller_matched_form,
                "brysbaert_biemiller_raw_aoa": aoa.brysbaert_biemiller_raw_aoa,
                "current_age_band": word_row["age_band"],
                "proposed_age_band": aoa.proposed_age_band,
                "review_status": "in_review",
                "review_notes": aoa.review_notes,
            }
        )

        recommendation_rows.append(
            {
                "word_key": word_row["word_key"],
                "normalised_word": word_row["normalised_word"],
                "display_word": word_row["display_word"],
                "dialect_code": word_row["dialect_code"],
                "current_frequency_band": word_row["frequency_band"],
                "recommended_frequency_band": frequency.proposed_band,
                "frequency_match_status": frequency.match_status,
                "frequency_evidence": ""
                if frequency.raw_frequency is None
                else f"BNC frequency={frequency.raw_frequency:g}; rank={frequency.rank}",
                "current_age_band": word_row["age_band"],
                "recommended_age_band": aoa.proposed_age_band,
                "aoa_evidence_status": aoa.source_data_status,
                "recommendation_status": (
                    "frequency_ready_aoa_pending"
                    if frequency.proposed_band and not aoa.proposed_age_band
                    else "frequency_and_aoa_ready"
                    if frequency.proposed_band and aoa.proposed_age_band
                    else "needs_source_review"
                ),
                "review_notes": "Audit recommendation only; do not update active canonical_words.csv until reviewed.",
            }
        )

    update_source_register()
    write_csv(BRITISH_FREQUENCY_INTAKE_PATH, list(frequency_rows[0].keys()), frequency_rows)
    write_csv(AOA_INTAKE_PATH, list(aoa_rows[0].keys()), aoa_rows)
    write_csv(RECOMMENDATIONS_PATH, list(recommendation_rows[0].keys()), recommendation_rows)

    _source_fields, source_rows = read_csv(SOURCES_PATH)
    summary = {
        "generated_at": date.today().isoformat(),
        "rows": len(word_rows),
        "unique_word_keys": len({row["word_key"] for row in word_rows}),
        "row_order_matches_canonical_words": True,
        "active_canonical_words_updated": False,
        "british_frequency_file": str(BRITISH_FREQUENCY_INTAKE_PATH.relative_to(Path.cwd())),
        "aoa_file": str(AOA_INTAKE_PATH.relative_to(Path.cwd())),
        "recommendations_file": str(RECOMMENDATIONS_PATH.relative_to(Path.cwd())),
        "bnc_source_cache": str(BNC_SOURCE_CACHE.relative_to(Path.cwd())),
        "frequency_match_counts": dict(frequency_counts),
        "proposed_frequency_band_counts": dict(proposed_frequency_counts),
        "aoa_source_data_status": {
            "kuperman": kuperman_status,
            "brysbaert_biemiller": bb_status,
        },
        "proposed_age_band_counts": dict(proposed_age_counts),
        "source_rows_after_update": len(source_rows),
        "source_importability_status": {
            BNC_SOURCE_KEY: "requires_legal_review",
            KUPERMAN_SOURCE_KEY: "requires_legal_review",
            BRYSBAERT_BIEMILLER_SOURCE_KEY: "requires_legal_review",
        },
        "hard_boundaries": [
            "no canonical_words.csv band update",
            "no Supabase import",
            "no database mutation",
            "no migrations",
            "no importer changes",
            "no runtime hooks",
            "no resolver changes",
            "no assignment changes",
            "no evidence/proficiency changes",
            "no Word Treasure changes",
        ],
    }
    SUMMARY_PATH.write_text(json.dumps(summary, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    update_build_summary(summary)
    print(json.dumps(summary, indent=2, sort_keys=True))


if __name__ == "__main__":
    main()

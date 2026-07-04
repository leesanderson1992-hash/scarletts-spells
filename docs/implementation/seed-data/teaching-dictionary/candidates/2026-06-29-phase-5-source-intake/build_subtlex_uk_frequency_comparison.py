#!/usr/bin/env python3
"""Compare existing BNC frequency audit rows with SUBTLEX-UK/wordfreq.

This script is audit-only. It does not update canonical_words.csv, import data,
touch Supabase, or change runtime behavior.

Supply a local SUBTLEX-UK source file with:
- SUBTLEX_UK_PATH=/path/to/subtlex-uk.csv|tsv|txt|xlsx

If SUBTLEX-UK is not supplied, the script uses Python package `wordfreq` as the
primary comparison signal when it is installed/importable. This still remains
audit-only.
"""

from __future__ import annotations

import csv
import json
import os
import re
from collections import Counter
from datetime import date
from pathlib import Path
from typing import Any


BASE = Path(__file__).resolve().parent
CSV_DIR = BASE / "csv"

WORDS_PATH = CSV_DIR / "canonical_words.csv"
SOURCES_PATH = CSV_DIR / "teaching_content_sources.csv"
BUILD_SUMMARY_PATH = BASE / "build_summary.json"
BNC_INTAKE_PATH = BASE / "british_frequency_intake_all_canonical_words.csv"

COMPARISON_PATH = BASE / "bnc_subtlex_uk_frequency_comparison_all_canonical_words.csv"
SUMMARY_PATH = BASE / "bnc_subtlex_uk_frequency_comparison_summary.json"

SUBTLEX_SOURCE_KEY = "subtlex_uk_van_heuven_2014_reference_2026_07_01"
SUBTLEX_SOURCE_NAME = "SUBTLEX-UK"
SUBTLEX_SOURCE_URL = "https://doi.org/10.1080/17470218.2013.850521"
SUBTLEX_SOURCE_LICENCE = (
    "Published dataset/source terms not verified in this pass; legal review required before product import"
)
SUBTLEX_SOURCE_NOTE = (
    "British subtitle-based word frequency reference for audit comparison with "
    "BNC. Source data file was not bundled by this pass unless SUBTLEX_UK_PATH "
    "is supplied. Active canonical_words.csv frequency_band values are not "
    "updated by this pass."
)

WORDFREQ_SOURCE_KEY = "wordfreq_python_package_fallback_2026_07_01"
WORDFREQ_SOURCE_NAME = "wordfreq Python package"
WORDFREQ_SOURCE_URL = "https://github.com/rspeer/wordfreq/"
WORDFREQ_SOURCE_LICENCE = "Apache-2.0 package licence; underlying data-source importability requires review"
WORDFREQ_SOURCE_NOTE = (
    "Primary frequency audit signal used when SUBTLEX-UK source data is not "
    "locally supplied. BNC is retained as fallback/comparison evidence. Active "
    "canonical_words.csv frequency_band values are not updated by this pass."
)


def read_csv(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    with path.open(newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        return list(reader.fieldnames or []), list(reader)


def write_csv(path: Path, fieldnames: list[str], rows: list[dict[str, str]]) -> None:
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def normalize(value: str) -> str:
    return value.strip().lower().replace("’", "'")


def split_components(value: str) -> list[str]:
    return [piece for piece in re.split(r"[/\s-]+", normalize(value)) if piece]


def find_column(headers: list[str], candidates: list[str]) -> str | None:
    lowered = {header.lower().strip(): header for header in headers}
    for candidate in candidates:
        if candidate.lower() in lowered:
            return lowered[candidate.lower()]
    for header in headers:
        normal = re.sub(r"[^a-z0-9]", "", header.lower())
        for candidate in candidates:
            if re.sub(r"[^a-z0-9]", "", candidate.lower()) == normal:
                return header
    return None


def load_table(path: Path) -> tuple[list[str], list[dict[str, Any]]]:
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        try:
            from openpyxl import load_workbook
        except ModuleNotFoundError as exc:
            raise SystemExit("openpyxl is required to read SUBTLEX-UK XLSX files") from exc
        workbook = load_workbook(path, read_only=True, data_only=True)
        worksheet = workbook[workbook.sheetnames[0]]
        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            return [], []
        headers = [str(value).strip() if value is not None else "" for value in rows[0]]
        return headers, [
            {headers[i]: value for i, value in enumerate(row) if i < len(headers)}
            for row in rows[1:]
        ]

    delimiter = "\t" if path.suffix.lower() in {".tsv", ".txt"} else ","
    with path.open(newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f, delimiter=delimiter)
        return list(reader.fieldnames or []), list(reader)


def parse_float(value: Any) -> float | None:
    if value is None:
        return None
    text = str(value).strip().replace(",", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def band_from_subtlex(zipf: float | None, frequency: float | None) -> str:
    if zipf is not None:
        if zipf >= 4.0:
            return "high"
        if zipf >= 3.0:
            return "medium"
        return "low"
    if frequency is not None:
        if frequency >= 100:
            return "high"
        if frequency >= 10:
            return "medium"
        return "low"
    return ""


def band_from_zipf(zipf: float | None) -> str:
    if zipf is None:
        return ""
    if zipf >= 4.0:
        return "high"
    if zipf >= 3.0:
        return "medium"
    return "low"


def load_subtlex(path: Path | None) -> tuple[str, dict[str, dict[str, str]]]:
    if not path:
        return "source_file_not_supplied", {}
    if not path.exists():
        return "source_file_missing", {}

    headers, rows = load_table(path)
    if not rows:
        return "source_file_empty", {}

    word_col = find_column(headers, ["word", "Word", "Spelling", "spelling", "Wordform"])
    freq_col = find_column(
        headers,
        [
            "FREQcount",
            "FreqCount",
            "frequency",
            "Frequency",
            "Freq",
            "SUBTLEX",
            "SUBTLWF",
        ],
    )
    zipf_col = find_column(headers, ["Zipf", "Zipf-value", "ZipfSUBTLEX", "Zipf frequency"])
    cd_col = find_column(headers, ["CDcount", "ContextualDiversity", "CD", "CDlow"])

    if not word_col or (not freq_col and not zipf_col):
        return "source_file_unrecognized_columns", {}

    data: dict[str, dict[str, str]] = {}
    for row in rows:
        word = normalize(str(row.get(word_col, "") or ""))
        if not word or word in data:
            continue
        raw_freq = parse_float(row.get(freq_col)) if freq_col else None
        raw_zipf = parse_float(row.get(zipf_col)) if zipf_col else None
        data[word] = {
            "matched_form": word,
            "raw_frequency": "" if raw_freq is None else f"{raw_freq:g}",
            "zipf": "" if raw_zipf is None else f"{raw_zipf:g}",
            "contextual_diversity": ""
            if not cd_col or row.get(cd_col) is None
            else str(row.get(cd_col)).strip(),
            "proposed_band": band_from_subtlex(raw_zipf, raw_freq),
        }
    return "loaded", data


def load_wordfreq() -> tuple[str, Any]:
    try:
        from wordfreq import zipf_frequency
    except ModuleNotFoundError:
        return "not_installed", None
    return "loaded", zipf_frequency


def select_subtlex(word: str, subtlex: dict[str, dict[str, str]]) -> dict[str, str]:
    lookup = normalize(word)
    if lookup in subtlex:
        return {"match_status": "exact", **subtlex[lookup]}

    components = split_components(word)
    if len(components) > 1 and all(component in subtlex for component in components):
        freqs = [parse_float(subtlex[component].get("raw_frequency")) for component in components]
        zipfs = [parse_float(subtlex[component].get("zipf")) for component in components]
        numeric_freqs = [value for value in freqs if value is not None]
        numeric_zipfs = [value for value in zipfs if value is not None]
        total_freq = sum(numeric_freqs) if numeric_freqs else None
        max_zipf = max(numeric_zipfs) if numeric_zipfs else None
        return {
            "match_status": "component",
            "matched_form": " | ".join(components),
            "raw_frequency": "" if total_freq is None else f"{total_freq:g}",
            "zipf": "" if max_zipf is None else f"{max_zipf:g}",
            "contextual_diversity": "",
            "proposed_band": band_from_subtlex(max_zipf, total_freq),
        }

    return {
        "match_status": "unmatched",
        "matched_form": "",
        "raw_frequency": "",
        "zipf": "",
        "contextual_diversity": "",
        "proposed_band": "",
    }


def select_wordfreq(word: str, zipf_frequency: Any) -> dict[str, str]:
    if not zipf_frequency:
        return {
            "match_status": "unavailable",
            "matched_form": "",
            "zipf": "",
            "proposed_band": "",
        }
    lookup = normalize(word)
    zipf = float(zipf_frequency(lookup, "en"))
    if zipf > 0:
        return {
            "match_status": "exact",
            "matched_form": lookup,
            "zipf": f"{zipf:g}",
            "proposed_band": band_from_zipf(zipf),
        }

    components = split_components(word)
    if len(components) > 1:
        zipfs = [float(zipf_frequency(component, "en")) for component in components]
        if all(value > 0 for value in zipfs):
            max_zipf = max(zipfs)
            return {
                "match_status": "component",
                "matched_form": " | ".join(components),
                "zipf": f"{max_zipf:g}",
                "proposed_band": band_from_zipf(max_zipf),
            }

    return {
        "match_status": "unmatched",
        "matched_form": "",
        "zipf": "",
        "proposed_band": "",
    }


def compare_bands(bnc_band: str, comparison_band: str, source_status: str) -> str:
    if source_status not in {"subtlex_loaded", "wordfreq_loaded"}:
        return "comparison_source_pending"
    if not bnc_band and not comparison_band:
        return "both_missing"
    if bnc_band and not comparison_band:
        return "bnc_only"
    if comparison_band and not bnc_band:
        return "comparison_only"
    if bnc_band == comparison_band:
        return "agree"
    return "disagree"


def update_source_register() -> None:
    fieldnames, rows = read_csv(SOURCES_PATH)
    replacements = {
        SUBTLEX_SOURCE_KEY: {
            "source_key": SUBTLEX_SOURCE_KEY,
            "source_category": "reference_only",
            "source_name": SUBTLEX_SOURCE_NAME,
            "source_url": SUBTLEX_SOURCE_URL,
            "source_licence": SUBTLEX_SOURCE_LICENCE,
            "source_use_note": SUBTLEX_SOURCE_NOTE,
            "importability_status": "requires_legal_review",
            "legal_review_status": "required",
        },
        WORDFREQ_SOURCE_KEY: {
            "source_key": WORDFREQ_SOURCE_KEY,
            "source_category": "open_licensed",
            "source_name": WORDFREQ_SOURCE_NAME,
            "source_url": WORDFREQ_SOURCE_URL,
            "source_licence": WORDFREQ_SOURCE_LICENCE,
            "source_use_note": WORDFREQ_SOURCE_NOTE,
            "importability_status": "requires_legal_review",
            "legal_review_status": "required",
        },
    }
    rows = [row for row in rows if row.get("source_key") not in replacements]
    rows.extend(replacements.values())
    write_csv(SOURCES_PATH, fieldnames, rows)


def update_build_summary(summary: dict[str, Any]) -> None:
    if not BUILD_SUMMARY_PATH.exists():
        return
    data = json.loads(BUILD_SUMMARY_PATH.read_text(encoding="utf-8"))
    data.update(
        {
            "subtlex_uk_frequency_comparison_created": True,
            "subtlex_uk_frequency_comparison_rows": summary["rows"],
            "subtlex_uk_source_data_status": summary["subtlex_source_data_status"],
            "subtlex_uk_match_counts": summary["subtlex_match_counts"],
            "wordfreq_fallback_source_data_status": summary["wordfreq_source_data_status"],
            "wordfreq_fallback_match_counts": summary["wordfreq_match_counts"],
            "frequency_comparison_source_used": summary["comparison_source_used"],
            "bnc_subtlex_band_comparison_counts": summary["band_comparison_counts"],
            "canonical_words_frequency_band_values_unchanged_after_subtlex_comparison": True,
            "sources": summary["source_rows_after_update"],
        }
    )
    BUILD_SUMMARY_PATH.write_text(
        json.dumps(data, indent=2, sort_keys=True) + "\n", encoding="utf-8"
    )


def main() -> None:
    _word_fields, word_rows = read_csv(WORDS_PATH)
    _bnc_fields, bnc_rows = read_csv(BNC_INTAKE_PATH)
    if len(word_rows) != 874 or len(bnc_rows) != 874:
        raise SystemExit("Expected 874 canonical and BNC audit rows")
    if [row["word_key"] for row in word_rows] != [row["word_key"] for row in bnc_rows]:
        raise SystemExit("BNC audit row order does not match canonical_words.csv")

    subtlex_path = Path(os.environ["SUBTLEX_UK_PATH"]) if os.environ.get("SUBTLEX_UK_PATH") else None
    subtlex_status, subtlex = load_subtlex(subtlex_path)
    wordfreq_status, zipf_frequency = load_wordfreq()
    comparison_source_used = "subtlex_uk" if subtlex_status == "loaded" else "wordfreq"
    comparison_status = (
        "subtlex_loaded"
        if subtlex_status == "loaded"
        else "wordfreq_loaded"
        if wordfreq_status == "loaded"
        else "source_pending"
    )

    rows: list[dict[str, str]] = []
    match_counts: Counter[str] = Counter()
    wordfreq_counts: Counter[str] = Counter()
    comparison_counts: Counter[str] = Counter()

    for word_row, bnc_row in zip(word_rows, bnc_rows):
        subtlex_row = select_subtlex(word_row["normalised_word"], subtlex) if subtlex else {
            "match_status": "unavailable" if subtlex_status != "loaded" else "unmatched",
            "matched_form": "",
            "raw_frequency": "",
            "zipf": "",
            "contextual_diversity": "",
            "proposed_band": "",
        }
        wordfreq_row = select_wordfreq(word_row["normalised_word"], zipf_frequency)
        comparison_band = (
            subtlex_row.get("proposed_band", "")
            if subtlex_status == "loaded"
            else wordfreq_row.get("proposed_band", "")
        )
        comparison = compare_bands(
            bnc_row.get("proposed_frequency_band", ""),
            comparison_band,
            comparison_status,
        )
        match_counts[subtlex_row["match_status"]] += 1
        wordfreq_counts[wordfreq_row["match_status"]] += 1
        comparison_counts[comparison] += 1
        rows.append(
            {
                "word_key": word_row["word_key"],
                "normalised_word": word_row["normalised_word"],
                "display_word": word_row["display_word"],
                "dialect_code": word_row["dialect_code"],
                "bnc_match_status": bnc_row.get("match_status", ""),
                "bnc_matched_form": bnc_row.get("matched_form", ""),
                "bnc_raw_frequency_per_million": bnc_row.get("raw_frequency_per_million", ""),
                "bnc_frequency_rank": bnc_row.get("frequency_rank", ""),
                "bnc_proposed_frequency_band": bnc_row.get("proposed_frequency_band", ""),
                "subtlex_source_data_status": subtlex_status,
                "subtlex_source_name": SUBTLEX_SOURCE_NAME,
                "subtlex_source_url": SUBTLEX_SOURCE_URL,
                "subtlex_match_status": subtlex_row.get("match_status", ""),
                "subtlex_matched_form": subtlex_row.get("matched_form", ""),
                "subtlex_raw_frequency": subtlex_row.get("raw_frequency", ""),
                "subtlex_zipf": subtlex_row.get("zipf", ""),
                "subtlex_contextual_diversity": subtlex_row.get("contextual_diversity", ""),
                "subtlex_proposed_frequency_band": subtlex_row.get("proposed_band", ""),
                "wordfreq_source_data_status": wordfreq_status,
                "wordfreq_source_name": WORDFREQ_SOURCE_NAME,
                "wordfreq_source_url": WORDFREQ_SOURCE_URL,
                "wordfreq_match_status": wordfreq_row.get("match_status", ""),
                "wordfreq_matched_form": wordfreq_row.get("matched_form", ""),
                "wordfreq_zipf": wordfreq_row.get("zipf", ""),
                "wordfreq_proposed_frequency_band": wordfreq_row.get("proposed_band", ""),
                "comparison_source_used": comparison_source_used,
                "comparison_proposed_frequency_band": comparison_band,
                "band_comparison_status": comparison,
                "review_status": "in_review",
                "review_notes": (
                    "SUBTLEX-UK source file not supplied; wordfreq used as primary audit comparison signal."
                    if subtlex_status != "loaded" and wordfreq_status == "loaded"
                    else "SUBTLEX-UK source file not supplied and wordfreq is not installed; comparison pending."
                    if subtlex_status != "loaded"
                    else "SUBTLEX-UK audit comparison only; review before updating active frequency_band."
                ),
            }
        )

    update_source_register()
    write_csv(COMPARISON_PATH, list(rows[0].keys()), rows)
    _source_fields, source_rows = read_csv(SOURCES_PATH)
    summary = {
        "generated_at": date.today().isoformat(),
        "rows": len(rows),
        "unique_word_keys": len({row["word_key"] for row in rows}),
        "row_order_matches_canonical_words": True,
        "active_canonical_words_updated": False,
        "comparison_file": str(COMPARISON_PATH.relative_to(Path.cwd())),
        "subtlex_source_data_status": subtlex_status,
        "subtlex_path": "" if subtlex_path is None else str(subtlex_path),
        "wordfreq_source_data_status": wordfreq_status,
        "comparison_source_used": comparison_source_used,
        "subtlex_match_counts": dict(match_counts),
        "wordfreq_match_counts": dict(wordfreq_counts),
        "band_comparison_counts": dict(comparison_counts),
        "source_rows_after_update": len(source_rows),
        "source_importability_status": {
            SUBTLEX_SOURCE_KEY: "requires_legal_review",
            WORDFREQ_SOURCE_KEY: "requires_legal_review",
        },
        "hard_boundaries": [
            "no canonical_words.csv band update",
            "no Supabase import",
            "no database mutation",
            "no migrations",
            "no importer changes",
            "no runtime hooks",
            "no resolver changes",
            "no assignment changes",
            "no evidence/proficiency changes",
            "no Word Treasure changes",
        ],
    }
    SUMMARY_PATH.write_text(json.dumps(summary, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    update_build_summary(summary)
    print(json.dumps(summary, indent=2, sort_keys=True))


if __name__ == "__main__":
    main()

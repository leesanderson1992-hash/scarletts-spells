#!/usr/bin/env python3
"""Apply manual age fill-ins to age recommendation audit artifacts.

This script is audit-only. It updates candidate recommendation CSVs and audit
summaries, but does not update canonical_words.csv, import data, touch
Supabase, or change runtime behavior.
"""

from __future__ import annotations

import csv
import hashlib
import json
from collections import Counter
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


BASE = Path(__file__).resolve().parent
CSV_DIR = BASE / "csv"

WORDS_PATH = CSV_DIR / "canonical_words.csv"
SOURCES_PATH = CSV_DIR / "teaching_content_sources.csv"
AGE_RECOMMENDATIONS_PATH = BASE / "canonical_words_age_recommendations.csv"
FREQUENCY_RECOMMENDATIONS_PATH = BASE / "canonical_words_frequency_aoa_band_recommendations.csv"
AGE_SUMMARY_PATH = BASE / "age_recommendation_population_summary.json"
BUILD_SUMMARY_PATH = BASE / "build_summary.json"

FILL_SOURCE_PATH = BASE / "missing_age_evidence_populated_uk_curriculum.xlsx"
APPLIED_NOTES_PATH = BASE / "missing_age_evidence_fill_in_applied_notes.csv"
SUMMARY_PATH = BASE / "missing_age_evidence_fill_in_applied_summary.json"

SOURCE_KEY = "katie_missing_age_evidence_uk_curriculum_fill_2026_07_02"
SOURCE_NAME = "Katie Sanderson missing age evidence UK curriculum fill"
SOURCE_URL = (
    "docs/implementation/seed-data/teaching-dictionary/candidates/"
    "2026-06-29-phase-5-source-intake/missing_age_evidence_populated_uk_curriculum.xlsx"
)
SOURCE_NOTE = (
    "Human-reviewed manual fill for words with no test-based AoA or UK age "
    "estimate match. Uses UK curriculum pattern placement and a project "
    "year-to-age mapping; candidate audit only, not active canonical_words.csv "
    "age truth."
)


def read_csv(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    with path.open(newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        return list(reader.fieldnames or []), list(reader)


def write_csv(path: Path, fieldnames: list[str], rows: list[dict[str, str]]) -> None:
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def age_band(age: int | None) -> str:
    if age is None:
        return ""
    if age <= 7:
        return "early_primary"
    if age <= 9:
        return "middle_primary"
    if age <= 11:
        return "upper_primary"
    if age <= 13:
        return "lower_secondary"
    if age <= 15:
        return "mid_secondary"
    return "later_review"


def parse_age(value: Any) -> int | None:
    text = str(value or "").strip()
    if not text:
        return None
    try:
        return int(float(text))
    except ValueError as exc:
        raise SystemExit(f"Invalid suggested_age_number value: {value!r}") from exc


def load_fill_rows() -> dict[str, dict[str, str]]:
    workbook = load_workbook(FILL_SOURCE_PATH, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = sheet.iter_rows(values_only=True)
    headers = [str(value).strip() if value is not None else "" for value in next(rows)]
    required = [
        "word_key",
        "normalised_word",
        "display_word",
        "dialect_code",
        "suggested_age_number",
        "suggested_age_band",
        "source_category",
        "source_name",
        "source_url_or_path",
        "source_use_note",
        "confidence",
        "review_status",
        "review_notes",
    ]
    missing = [header for header in required if header not in headers]
    if missing:
        raise SystemExit(f"Fill workbook missing required columns: {missing}")
    index = {header: idx for idx, header in enumerate(headers)}
    result: dict[str, dict[str, str]] = {}
    for values in rows:
        if not values or not values[index["word_key"]]:
            continue
        row = {
            header: "" if values[index[header]] is None else str(values[index[header]]).strip()
            for header in headers
            if header
        }
        age = parse_age(row.get("suggested_age_number"))
        if age is None:
            raise SystemExit(f"Missing suggested_age_number for {row.get('word_key')}")
        if row["word_key"] in result:
            raise SystemExit(f"Duplicate fill row for {row['word_key']}")
        result[row["word_key"]] = row
    return result


def update_source_register() -> None:
    fieldnames, rows = read_csv(SOURCES_PATH)
    replacement = {
        "source_key": SOURCE_KEY,
        "source_category": "internal_authored",
        "source_name": SOURCE_NAME,
        "source_url": SOURCE_URL,
        "source_licence": "internal/project-authored manual review",
        "source_use_note": SOURCE_NOTE,
        "importability_status": "importable",
        "legal_review_status": "passed",
    }
    rows = [row for row in rows if row.get("source_key") != SOURCE_KEY]
    rows.append(replacement)
    write_csv(SOURCES_PATH, fieldnames, rows)


def ensure_fields(fieldnames: list[str], additions: list[str]) -> list[str]:
    result = list(fieldnames)
    for field in additions:
        if field not in result:
            result.append(field)
    return result


def update_json(path: Path, updates: dict[str, Any]) -> None:
    if not path.exists():
        return
    data = json.loads(path.read_text(encoding="utf-8"))
    data.update(updates)
    path.write_text(json.dumps(data, indent=2, sort_keys=True) + "\n", encoding="utf-8")


def main() -> None:
    if not FILL_SOURCE_PATH.exists():
        raise SystemExit(f"Fill workbook missing: {FILL_SOURCE_PATH}")

    _word_fields, word_rows = read_csv(WORDS_PATH)
    age_fields, age_rows = read_csv(AGE_RECOMMENDATIONS_PATH)
    frequency_fields, frequency_rows = read_csv(FREQUENCY_RECOMMENDATIONS_PATH)
    if [row["word_key"] for row in word_rows] != [row["word_key"] for row in age_rows]:
        raise SystemExit("Age recommendation rows do not match canonical word order")
    if [row["word_key"] for row in word_rows] != [row["word_key"] for row in frequency_rows]:
        raise SystemExit("Frequency recommendation rows do not match canonical word order")

    fill_rows = load_fill_rows()
    missing_keys = {
        row["word_key"]
        for row in age_rows
        if row.get("age_source_status") == "missing_age_evidence"
    }
    if set(fill_rows) != missing_keys:
        missing = sorted(missing_keys - set(fill_rows))
        extra = sorted(set(fill_rows) - missing_keys)
        raise SystemExit(f"Fill rows do not match missing age keys; missing={missing}, extra={extra}")

    applied_notes: list[dict[str, str]] = []
    applied = 0
    for age_row, frequency_row in zip(age_rows, frequency_rows):
        key = age_row["word_key"]
        fill = fill_rows.get(key)
        if not fill:
            continue
        suggested_age = parse_age(fill["suggested_age_number"])
        band = age_band(suggested_age)
        flags = [
            flag
            for flag in age_row.get("review_flags", "").split("|")
            if flag and flag != "missing_age_evidence"
        ]
        flags.append("manual_uk_curriculum_age_fill")
        if fill.get("suggested_age_band"):
            flags.append(f"source_band:{fill['suggested_age_band']}")
        review_flags = "|".join(flags)

        age_row.update(
            {
                "recommended_age_number_unrounded": str(suggested_age),
                "recommended_age_number": str(suggested_age),
                "recommended_age_band": band,
                "age_source_status": "manual_uk_curriculum_fill",
                "uk_match_status": "manual_estimate",
                "uk_curriculum_age": str(suggested_age),
                "review_flags": review_flags,
                "recommendation_status": "age_ready_with_review_flags",
                "source_category": "internal_authored",
                "source_name": SOURCE_NAME,
                "source_url": SOURCE_URL,
                "source_licence": "see source register rows",
                "source_use_note": SOURCE_NOTE,
                "review_status": fill.get("review_status") or "in_review",
                "review_notes": (
                    f"Manual UK curriculum fill applied. Source band="
                    f"{fill.get('suggested_age_band', '')}. "
                    f"{fill.get('review_notes', '')}"
                ).strip(),
                "manual_fill_source_name": fill.get("source_name", ""),
                "manual_fill_source_url_or_path": fill.get("source_url_or_path", ""),
                "manual_fill_confidence": fill.get("confidence", ""),
                "manual_fill_review_notes": fill.get("review_notes", ""),
            }
        )
        frequency_row.update(
            {
                "recommended_age_number_unrounded": str(suggested_age),
                "recommended_age_number": str(suggested_age),
                "recommended_age_band": band,
                "age_source_status": "manual_uk_curriculum_fill",
                "aoa_evidence_status": "manual_uk_curriculum_fill",
                "uk_curriculum_age": str(suggested_age),
                "age_review_flags": review_flags,
            }
        )
        if frequency_row.get("recommended_frequency_band"):
            frequency_row["recommendation_status"] = "frequency_and_age_ready_with_review_flags"
        else:
            frequency_row["recommendation_status"] = "age_ready_frequency_pending"
        frequency_row["review_notes"] = (
            f"{frequency_row.get('review_notes', '')} Manual missing-age fill "
            f"applied from {SOURCE_NAME}; active canonical_words.csv age_band unchanged."
        ).strip()

        applied += 1
        applied_notes.append(
            {
                "word_key": key,
                "normalised_word": age_row["normalised_word"],
                "suggested_age_number": str(suggested_age),
                "recommended_age_band": band,
                "suggested_age_band_original": fill.get("suggested_age_band", ""),
                "confidence": fill.get("confidence", ""),
                "review_status": fill.get("review_status", ""),
                "review_notes": fill.get("review_notes", ""),
            }
        )

    update_source_register()
    age_fields = ensure_fields(
        age_fields,
        [
            "manual_fill_source_name",
            "manual_fill_source_url_or_path",
            "manual_fill_confidence",
            "manual_fill_review_notes",
        ],
    )
    frequency_fields = ensure_fields(
        frequency_fields,
        [
            "recommended_age_number_unrounded",
            "recommended_age_number",
            "age_source_status",
            "aoa_median_age",
            "aoa_meaning_row_count",
            "uk_curriculum_age",
            "age_review_flags",
        ],
    )
    write_csv(AGE_RECOMMENDATIONS_PATH, age_fields, age_rows)
    write_csv(FREQUENCY_RECOMMENDATIONS_PATH, frequency_fields, frequency_rows)
    write_csv(
        APPLIED_NOTES_PATH,
        [
            "word_key",
            "normalised_word",
            "suggested_age_number",
            "recommended_age_band",
            "suggested_age_band_original",
            "confidence",
            "review_status",
            "review_notes",
        ],
        applied_notes,
    )

    _source_fields, source_rows = read_csv(SOURCES_PATH)
    source_counts = Counter(row.get("age_source_status", "") for row in age_rows)
    band_counts = Counter(row.get("recommended_age_band", "") or "blank" for row in age_rows)
    flag_counts: Counter[str] = Counter()
    for row in age_rows:
        for flag in row.get("review_flags", "").split("|"):
            if flag:
                flag_counts[flag] += 1

    summary = {
        "generated_at": date.today().isoformat(),
        "rows_applied": applied,
        "fill_source_file": str(FILL_SOURCE_PATH.relative_to(Path.cwd())),
        "fill_source_sha256": sha256(FILL_SOURCE_PATH),
        "applied_notes_file": str(APPLIED_NOTES_PATH.relative_to(Path.cwd())),
        "active_canonical_words_updated": False,
        "age_source_status_counts_after_apply": dict(source_counts),
        "recommended_age_band_counts_after_apply": dict(band_counts),
        "review_flag_counts_after_apply": dict(flag_counts),
        "source_rows_after_update": len(source_rows),
        "source_importability_status": {SOURCE_KEY: "importable"},
        "legal_review_status": {SOURCE_KEY: "passed"},
        "hard_boundaries": [
            "no canonical_words.csv age_band update",
            "no Supabase import",
            "no database mutation",
            "no migrations",
            "no importer changes",
            "no runtime hooks",
            "no resolver changes",
            "no assignment changes",
            "no evidence/proficiency changes",
            "no Word Treasure changes",
        ],
    }
    SUMMARY_PATH.write_text(json.dumps(summary, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    update_json(
        AGE_SUMMARY_PATH,
        {
            "missing_age_fill_in_applied": True,
            "missing_age_fill_in_rows_applied": applied,
            "missing_age_fill_in_summary_file": str(SUMMARY_PATH.relative_to(Path.cwd())),
            "age_source_status_counts_after_fill": dict(source_counts),
            "recommended_age_band_counts_after_fill": dict(band_counts),
            "active_canonical_words_updated": False,
        },
    )
    update_json(
        BUILD_SUMMARY_PATH,
        {
            "missing_age_fill_in_applied": True,
            "missing_age_fill_in_rows_applied": applied,
            "canonical_words_age_band_values_unchanged": True,
            "recommended_age_band_counts_after_fill": dict(band_counts),
            "sources": len(source_rows),
        },
    )
    print(json.dumps(summary, indent=2, sort_keys=True))


if __name__ == "__main__":
    main()

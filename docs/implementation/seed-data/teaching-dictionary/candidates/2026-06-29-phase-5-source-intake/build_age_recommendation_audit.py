#!/usr/bin/env python3
"""Build AoA + UK curriculum age recommendation audit artifacts.

This script is audit-only. It does not update canonical_words.csv, import data,
touch Supabase, or change runtime behavior.
"""

from __future__ import annotations

import csv
import hashlib
import json
import re
from collections import Counter, defaultdict
from datetime import date
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from statistics import median
from typing import Any

from openpyxl import load_workbook


BASE = Path(__file__).resolve().parent
CSV_DIR = BASE / "csv"

WORDS_PATH = CSV_DIR / "canonical_words.csv"
SOURCES_PATH = CSV_DIR / "teaching_content_sources.csv"
BUILD_SUMMARY_PATH = BASE / "build_summary.json"
FREQUENCY_RECOMMENDATIONS_PATH = BASE / "canonical_words_frequency_aoa_band_recommendations.csv"

AOA_SOURCE_PATH = BASE / "brysbaert_biemiller_test_based_aoa_master_source.xlsx"
UK_SOURCE_PATH = BASE / "uk_spelling_words_age_estimates.csv"

AOA_INTAKE_PATH = BASE / "aoa_test_based_intake_all_canonical_words.csv"
UK_INTAKE_PATH = BASE / "uk_curriculum_age_intake_all_canonical_words.csv"
AGE_RECOMMENDATIONS_PATH = BASE / "canonical_words_age_recommendations.csv"
SUMMARY_PATH = BASE / "age_recommendation_population_summary.json"

AOA_SOURCE_KEY = "brysbaert_biemiller_test_based_aoa_master_2026_07_02"
AOA_SOURCE_NAME = "Brysbaert-Biemiller test-based AoA master file"
AOA_SOURCE_URL = "docs/implementation/seed-data/teaching-dictionary/candidates/2026-06-29-phase-5-source-intake/brysbaert_biemiller_test_based_aoa_master_source.xlsx"
AOA_SOURCE_LICENCE = "Project legal/importability review passed for this candidate age-recommendation use"
AOA_SOURCE_NOTE = (
    "Legally approved test-based age-of-acquisition evidence. Uses AoAtestbased "
    "with median age across multiple meaning rows; candidate audit only, not "
    "active canonical_words.csv age truth."
)

UK_SOURCE_KEY = "uk_spelling_words_age_estimates_2026_07_02"
UK_SOURCE_NAME = "UK spelling words age estimates"
UK_SOURCE_URL = "docs/implementation/seed-data/teaching-dictionary/candidates/2026-06-29-phase-5-source-intake/uk_spelling_words_age_estimates.csv"
UK_SOURCE_LICENCE = "internal/project-authored age estimate extract"
UK_SOURCE_NOTE = (
    "Project-authored age estimate extract informed by UK curriculum spelling "
    "expectations. Used with AoA evidence for candidate age recommendations; "
    "active canonical_words.csv age_band values are not updated by this pass."
)


def read_csv(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    with path.open(newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        return list(reader.fieldnames or []), list(reader)


def write_csv(path: Path, fieldnames: list[str], rows: list[dict[str, str]]) -> None:
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def normalize(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().lower().replace("’", "'"))


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def aoa_code_to_age(code: Any) -> float | None:
    if code is None:
        return None
    try:
        number = int(float(code))
    except (TypeError, ValueError):
        return None
    return {
        2: 6.0,
        4: 8.5,
        6: 10.5,
        8: 12.5,
        10: 14.5,
        12: 16.5,
        13: 18.0,
        14: 18.0,
    }.get(number)


def fmt_number(value: float | None) -> str:
    if value is None:
        return ""
    if float(value).is_integer():
        return str(int(value))
    return f"{value:.2f}".rstrip("0").rstrip(".")


def round_half_up(value: float | None) -> int | None:
    if value is None:
        return None
    return int(Decimal(str(value)).quantize(Decimal("1"), rounding=ROUND_HALF_UP))


def age_band(age: int | None) -> str:
    if age is None:
        return ""
    if age <= 7:
        return "early_primary"
    if age <= 9:
        return "middle_primary"
    if age <= 11:
        return "upper_primary"
    if age <= 13:
        return "lower_secondary"
    if age <= 15:
        return "mid_secondary"
    return "later_review"


def load_aoa_source() -> dict[str, list[dict[str, Any]]]:
    workbook = load_workbook(AOA_SOURCE_PATH, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = sheet.iter_rows(values_only=True)
    headers = [str(value).strip() if value is not None else "" for value in next(rows)]
    index = {header: idx for idx, header in enumerate(headers)}
    required = {"WORD", "MEANING", "AoAtestbased", "AoArating", "LWV", "CDI", "Morr", "Floc"}
    missing = sorted(required - set(index))
    if missing:
        raise SystemExit(f"AoA workbook missing required columns: {missing}")

    by_word: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        word = normalize(row[index["WORD"]])
        if not word:
            continue
        code = row[index["AoAtestbased"]]
        converted_age = aoa_code_to_age(code)
        if converted_age is None:
            continue
        by_word[word].append(
            {
                "word": row[index["WORD"]],
                "meaning": "" if row[index["MEANING"]] is None else str(row[index["MEANING"]]),
                "aoa_code": code,
                "converted_age": converted_age,
                "aoa_rating": "" if row[index["AoArating"]] is None else str(row[index["AoArating"]]),
                "lwv": "" if row[index["LWV"]] is None else str(row[index["LWV"]]),
                "cdi": "" if row[index["CDI"]] is None else str(row[index["CDI"]]),
                "morr": "" if row[index["Morr"]] is None else str(row[index["Morr"]]),
                "floc": "" if row[index["Floc"]] is None else str(row[index["Floc"]]),
            }
        )
    return by_word


def load_uk_source() -> dict[str, dict[str, str]]:
    fieldnames, rows = read_csv(UK_SOURCE_PATH)
    if fieldnames != ["Word", "Age"]:
        raise SystemExit(f"UK age CSV must have headers Word,Age; got {fieldnames}")
    result: dict[str, dict[str, str]] = {}
    for row in rows:
        word = normalize(row.get("Word", ""))
        age_text = str(row.get("Age", "")).strip()
        if word and age_text:
            result[word] = {"word": row["Word"], "age": age_text}
    return result


def source_status(has_aoa: bool, has_uk: bool) -> str:
    if has_aoa and has_uk:
        return "aoa_and_uk"
    if has_aoa:
        return "aoa_only"
    if has_uk:
        return "uk_only"
    return "missing_age_evidence"


def recommendation_status(review_flags: list[str]) -> str:
    if "missing_age_evidence" in review_flags:
        return "needs_source_review"
    if review_flags:
        return "age_ready_with_review_flags"
    return "age_ready"


def update_source_register() -> None:
    fieldnames, rows = read_csv(SOURCES_PATH)
    replacements = {
        AOA_SOURCE_KEY: {
            "source_key": AOA_SOURCE_KEY,
            "source_category": "open_licensed",
            "source_name": AOA_SOURCE_NAME,
            "source_url": AOA_SOURCE_URL,
            "source_licence": AOA_SOURCE_LICENCE,
            "source_use_note": AOA_SOURCE_NOTE,
            "importability_status": "importable",
            "legal_review_status": "passed",
        },
        UK_SOURCE_KEY: {
            "source_key": UK_SOURCE_KEY,
            "source_category": "internal_authored",
            "source_name": UK_SOURCE_NAME,
            "source_url": UK_SOURCE_URL,
            "source_licence": UK_SOURCE_LICENCE,
            "source_use_note": UK_SOURCE_NOTE,
            "importability_status": "importable",
            "legal_review_status": "passed",
        },
    }
    rows = [row for row in rows if row.get("source_key") not in replacements]
    rows.extend(replacements.values())
    write_csv(SOURCES_PATH, fieldnames, rows)


def update_json(path: Path, updates: dict[str, Any]) -> None:
    if not path.exists():
        return
    data = json.loads(path.read_text(encoding="utf-8"))
    data.update(updates)
    path.write_text(json.dumps(data, indent=2, sort_keys=True) + "\n", encoding="utf-8")


def main() -> None:
    for path in [AOA_SOURCE_PATH, UK_SOURCE_PATH]:
        if not path.exists():
            raise SystemExit(f"Required source file missing: {path}")

    _word_fields, word_rows = read_csv(WORDS_PATH)
    if len(word_rows) != 874 or len({row["word_key"] for row in word_rows}) != 874:
        raise SystemExit("Expected 874 unique canonical word rows")

    _frequency_fields, frequency_rows = read_csv(FREQUENCY_RECOMMENDATIONS_PATH)
    if [row["word_key"] for row in word_rows] != [row["word_key"] for row in frequency_rows]:
        raise SystemExit("Frequency recommendation rows do not match canonical word order")

    aoa_by_word = load_aoa_source()
    uk_by_word = load_uk_source()

    aoa_rows: list[dict[str, str]] = []
    uk_rows: list[dict[str, str]] = []
    age_rows: list[dict[str, str]] = []
    updated_frequency_rows: list[dict[str, str]] = []

    match_counts: Counter[str] = Counter()
    recommended_band_counts: Counter[str] = Counter()
    review_flag_counts: Counter[str] = Counter()
    rounded_age_counts: Counter[str] = Counter()

    for word_row, frequency_row in zip(word_rows, frequency_rows):
        word_key = word_row["word_key"]
        word = word_row["normalised_word"]
        lookup = normalize(word)
        aoa_matches = aoa_by_word.get(lookup, [])
        uk_match = uk_by_word.get(lookup)

        converted_ages = [float(match["converted_age"]) for match in aoa_matches]
        median_aoa = float(median(converted_ages)) if converted_ages else None
        min_aoa = min(converted_ages) if converted_ages else None
        max_aoa = max(converted_ages) if converted_ages else None
        uk_age = float(uk_match["age"]) if uk_match else None

        final_unrounded: float | None
        if median_aoa is not None and uk_age is not None:
            final_unrounded = (median_aoa + uk_age) / 2
        elif median_aoa is not None:
            final_unrounded = median_aoa
        elif uk_age is not None:
            final_unrounded = uk_age
        else:
            final_unrounded = None

        final_age = round_half_up(final_unrounded)
        final_band = age_band(final_age)

        flags: list[str] = []
        if len(aoa_matches) > 1:
            flags.append("multiple_meanings_review")
        if median_aoa == 18.0 and min_aoa == 18.0:
            flags.append("aoa_18_plus_only")
        if median_aoa is not None and uk_age is not None and abs(median_aoa - uk_age) > 2:
            flags.append("aoa_uk_difference_over_2_years")
        if "/" in word or "-" in word or " " in word:
            flags.append("compound_or_multi_form_review")
        if final_age is None:
            flags.append("missing_age_evidence")

        status = source_status(bool(aoa_matches), bool(uk_match))
        match_counts[status] += 1
        recommended_band_counts[final_band or "blank"] += 1
        rounded_age_counts["blank" if final_age is None else str(final_age)] += 1
        for flag in flags:
            review_flag_counts[flag] += 1

        meaning_samples = "; ".join(match["meaning"] for match in aoa_matches[:5])
        aoa_codes = "|".join(str(match["aoa_code"]) for match in aoa_matches)
        converted_age_text = "|".join(fmt_number(match["converted_age"]) for match in aoa_matches)

        aoa_rows.append(
            {
                "word_key": word_key,
                "normalised_word": word,
                "display_word": word_row["display_word"],
                "dialect_code": word_row["dialect_code"],
                "source_category": "open_licensed",
                "source_name": AOA_SOURCE_NAME,
                "source_url": AOA_SOURCE_URL,
                "source_licence": AOA_SOURCE_LICENCE,
                "source_use_note": AOA_SOURCE_NOTE,
                "match_status": "exact" if aoa_matches else "unmatched",
                "matched_form": lookup if aoa_matches else "",
                "meaning_row_count": str(len(aoa_matches)),
                "aoa_codes": aoa_codes,
                "converted_age_values": converted_age_text,
                "median_aoa_age": fmt_number(median_aoa),
                "min_aoa_age": fmt_number(min_aoa),
                "max_aoa_age": fmt_number(max_aoa),
                "meaning_samples": meaning_samples,
                "review_status": "in_review",
                "review_notes": (
                    "Median AoA age calculated from AoAtestbased across matched meaning rows."
                    if aoa_matches
                    else "No exact AoA workbook match."
                ),
            }
        )

        uk_rows.append(
            {
                "word_key": word_key,
                "normalised_word": word,
                "display_word": word_row["display_word"],
                "dialect_code": word_row["dialect_code"],
                "source_category": "internal_authored",
                "source_name": UK_SOURCE_NAME,
                "source_url": UK_SOURCE_URL,
                "source_licence": UK_SOURCE_LICENCE,
                "source_use_note": UK_SOURCE_NOTE,
                "match_status": "exact" if uk_match else "unmatched",
                "matched_form": uk_match["word"] if uk_match else "",
                "uk_curriculum_age": "" if uk_age is None else str(int(uk_age)),
                "review_status": "in_review",
                "review_notes": (
                    "Exact UK curriculum age estimate match."
                    if uk_match
                    else "No exact UK curriculum age estimate match."
                ),
            }
        )

        age_row = {
            "word_key": word_key,
            "normalised_word": word,
            "display_word": word_row["display_word"],
            "dialect_code": word_row["dialect_code"],
            "current_age_band": word_row["age_band"],
            "aoa_match_status": "exact" if aoa_matches else "unmatched",
            "aoa_meaning_row_count": str(len(aoa_matches)),
            "aoa_median_age": fmt_number(median_aoa),
            "aoa_min_age": fmt_number(min_aoa),
            "aoa_max_age": fmt_number(max_aoa),
            "uk_match_status": "exact" if uk_match else "unmatched",
            "uk_curriculum_age": "" if uk_age is None else str(int(uk_age)),
            "age_source_status": status,
            "recommended_age_number_unrounded": fmt_number(final_unrounded),
            "recommended_age_number": "" if final_age is None else str(final_age),
            "recommended_age_band": final_band,
            "review_flags": "|".join(flags),
            "recommendation_status": recommendation_status(flags),
            "source_category": "internal_authored",
            "source_name": f"{AOA_SOURCE_NAME} + {UK_SOURCE_NAME}",
            "source_url": f"{AOA_SOURCE_URL}; {UK_SOURCE_URL}",
            "source_licence": "see source register rows",
            "source_use_note": (
                "Candidate age recommendation combines median test-based AoA "
                "and UK curriculum age estimate where both are available; "
                "active canonical_words.csv age_band is not updated."
            ),
            "review_status": "in_review",
            "review_notes": (
                "Half-up rounded from equal average of available age evidence. "
                f"Meaning samples: {meaning_samples[:500]}"
            ).strip(),
        }
        age_rows.append(age_row)

        updated = dict(frequency_row)
        updated.update(
            {
                "recommended_age_number_unrounded": age_row["recommended_age_number_unrounded"],
                "recommended_age_number": age_row["recommended_age_number"],
                "recommended_age_band": final_band,
                "age_source_status": status,
                "aoa_evidence_status": status,
                "aoa_median_age": age_row["aoa_median_age"],
                "aoa_meaning_row_count": age_row["aoa_meaning_row_count"],
                "uk_curriculum_age": age_row["uk_curriculum_age"],
                "age_review_flags": age_row["review_flags"],
            }
        )
        if updated.get("recommended_frequency_band") and final_band and flags:
            updated["recommendation_status"] = "frequency_and_age_ready_with_review_flags"
        elif updated.get("recommended_frequency_band") and final_band:
            updated["recommendation_status"] = "frequency_and_age_ready"
        elif updated.get("recommended_frequency_band"):
            updated["recommendation_status"] = "frequency_ready_age_pending"
        elif final_band:
            updated["recommendation_status"] = "age_ready_frequency_pending"
        else:
            updated["recommendation_status"] = "needs_source_review"
        updated["review_notes"] = (
            f"{frequency_row.get('review_notes', '')} Age recommendation audit: "
            "median AoA plus UK curriculum estimate where available; active "
            "canonical_words.csv age_band unchanged."
        ).strip()
        updated_frequency_rows.append(updated)

    update_source_register()
    write_csv(AOA_INTAKE_PATH, list(aoa_rows[0].keys()), aoa_rows)
    write_csv(UK_INTAKE_PATH, list(uk_rows[0].keys()), uk_rows)
    write_csv(AGE_RECOMMENDATIONS_PATH, list(age_rows[0].keys()), age_rows)
    write_csv(FREQUENCY_RECOMMENDATIONS_PATH, list(updated_frequency_rows[0].keys()), updated_frequency_rows)

    _source_fields, source_rows = read_csv(SOURCES_PATH)
    summary = {
        "generated_at": date.today().isoformat(),
        "rows": len(word_rows),
        "unique_word_keys": len({row["word_key"] for row in word_rows}),
        "row_order_matches_canonical_words": True,
        "active_canonical_words_updated": False,
        "aoa_source_file": str(AOA_SOURCE_PATH.relative_to(Path.cwd())),
        "aoa_source_sha256": sha256(AOA_SOURCE_PATH),
        "uk_source_file": str(UK_SOURCE_PATH.relative_to(Path.cwd())),
        "uk_source_sha256": sha256(UK_SOURCE_PATH),
        "aoa_intake_file": str(AOA_INTAKE_PATH.relative_to(Path.cwd())),
        "uk_intake_file": str(UK_INTAKE_PATH.relative_to(Path.cwd())),
        "age_recommendations_file": str(AGE_RECOMMENDATIONS_PATH.relative_to(Path.cwd())),
        "frequency_aoa_recommendations_file": str(FREQUENCY_RECOMMENDATIONS_PATH.relative_to(Path.cwd())),
        "age_source_status_counts": dict(match_counts),
        "recommended_age_band_counts": dict(recommended_band_counts),
        "recommended_age_number_counts": dict(rounded_age_counts),
        "review_flag_counts": dict(review_flag_counts),
        "source_rows_after_update": len(source_rows),
        "source_importability_status": {
            AOA_SOURCE_KEY: "importable",
            UK_SOURCE_KEY: "importable",
        },
        "legal_review_status": {
            AOA_SOURCE_KEY: "passed",
            UK_SOURCE_KEY: "passed",
        },
        "hard_boundaries": [
            "no canonical_words.csv age_band update",
            "no Supabase import",
            "no database mutation",
            "no migrations",
            "no importer changes",
            "no runtime hooks",
            "no resolver changes",
            "no assignment changes",
            "no evidence/proficiency changes",
            "no Word Treasure changes",
        ],
    }
    SUMMARY_PATH.write_text(json.dumps(summary, indent=2, sort_keys=True) + "\n", encoding="utf-8")

    update_json(
        BUILD_SUMMARY_PATH,
        {
            "age_recommendation_audit_created": True,
            "age_recommendation_audit_rows": len(age_rows),
            "age_source_status_counts": dict(match_counts),
            "recommended_age_band_counts": dict(recommended_band_counts),
            "canonical_words_age_band_values_unchanged": True,
            "sources": len(source_rows),
        },
    )
    print(json.dumps(summary, indent=2, sort_keys=True))


if __name__ == "__main__":
    main()

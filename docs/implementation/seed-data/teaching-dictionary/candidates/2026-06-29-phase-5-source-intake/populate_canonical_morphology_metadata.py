#!/usr/bin/env python3
"""Populate canonical morphology metadata from MorphoLex-en.

This script updates only the candidate CSV folder. It does not import data,
touch Supabase, mutate schema, or change runtime behavior.

Set MORPHOLEX_XLSX_PATH to the downloaded MorphoLEX_en.xlsx workbook if it is
not available at /private/tmp/MorphoLEX_en.xlsx.
"""

from __future__ import annotations

import csv
import json
import os
import re
import shutil
from collections import Counter
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except ModuleNotFoundError as exc:  # pragma: no cover - user environment guard.
    raise SystemExit(
        "openpyxl is required to read MorphoLEX_en.xlsx. Use the bundled "
        "workspace Python runtime or install openpyxl in your active Python."
    ) from exc


BASE = Path(__file__).resolve().parent
CSV_DIR = BASE / "csv"

METADATA_PATH = CSV_DIR / "canonical_word_metadata.csv"
WORDS_PATH = CSV_DIR / "canonical_words.csv"
SOURCES_PATH = CSV_DIR / "teaching_content_sources.csv"
BUILD_SUMMARY_PATH = BASE / "build_summary.json"

DEFAULT_MORPHOLEX_XLSX_PATH = Path("/private/tmp/MorphoLEX_en.xlsx")
MORPHOLEX_XLSX_PATH = Path(
    os.environ.get("MORPHOLEX_XLSX_PATH", str(DEFAULT_MORPHOLEX_XLSX_PATH))
)

BEFORE_PATH = BASE / "canonical_word_metadata_before_morphology_population.csv"
INTAKE_PATH = BASE / "morpholex_morphology_intake_all_canonical_words.csv"
COMPLEX_PATH = BASE / "morpholex_morphology_intake_unmatched_or_complex_rows.csv"
AUDIT_PATH = BASE / "canonical_word_metadata_morphology_population_audit.csv"
SUMMARY_PATH = BASE / "canonical_word_metadata_morphology_population_summary.json"

MORPHOLEX_SOURCE_KEY = "morpholex_en_2017_2026_07_01"
MORPHOLEX_SOURCE_NAME = "MorphoLex-en"
MORPHOLEX_SOURCE_URL = "https://github.com/hugomailhot/MorphoLex-en"
MORPHOLEX_SOURCE_LICENCE = "CC BY-NC-SA 4.0"
MORPHOLEX_SOURCE_NOTE = (
    "MorphoLex-en morphology segmentation and variables. Project legal/importability "
    "review passed for this candidate metadata use; no import performed by this pass. "
    "Citation: Sanchez-Gutierrez, Mailhot, Deacon et al. (2017), "
    "doi:10.3758/s13428-017-0981-8."
)

UNCHANGED_PRONUNCIATION_FIELDS = ["syllables", "phoneme_hint", "stress_pattern", "has_schwa"]
PRESERVED_FIELDS = ["grapheme_notes", "irregularity_notes"]


@dataclass(frozen=True)
class MorphEntry:
    word: str
    pos: str
    nmorph: str
    prs_signature: str
    segmentation: str
    sheet: str
    prefixes: tuple[str, ...]
    roots: tuple[str, ...]
    suffixes: tuple[str, ...]


def read_csv(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    with path.open(newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        return list(reader.fieldnames or []), list(reader)


def write_csv(path: Path, fieldnames: list[str], rows: list[dict[str, str]]) -> None:
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def lookup_key(value: str) -> str:
    return (
        value.strip()
        .lower()
        .replace("’", "'")
        .replace("`", "'")
    )


def collapsed_key(value: str) -> str:
    return re.sub(r"[^a-z0-9]", "", lookup_key(value))


def parse_parts(segmentation: str) -> tuple[tuple[str, ...], tuple[str, ...], tuple[str, ...]]:
    prefixes = tuple(re.findall(r"<([^<>{}()]+)<", segmentation))
    roots = tuple(re.findall(r"\(([^()]+)\)", segmentation))
    suffixes = tuple(re.findall(r">([^<>{}()]+)>", segmentation))
    return prefixes, roots, suffixes


def ordered_parts(entry: MorphEntry) -> list[str]:
    parts: list[str] = []
    for prefix in entry.prefixes:
        parts.append(f"prefix:{prefix}")
    for root in entry.roots:
        parts.append(f"root:{root}")
    for suffix in entry.suffixes:
        parts.append(f"suffix:{suffix}")
    return parts or [f"unsegmented:{entry.segmentation.strip('{}')}"]


def load_morpholex_entries(path: Path) -> dict[str, MorphEntry]:
    if not path.exists():
        raise SystemExit(
            f"MorphoLex workbook not found at {path}. Set MORPHOLEX_XLSX_PATH "
            "to the downloaded MorphoLEX_en.xlsx path."
        )

    workbook = load_workbook(path, read_only=True, data_only=True)
    entries: dict[str, MorphEntry] = {}
    for sheet_name in workbook.sheetnames:
        if sheet_name == "Presentation" or sheet_name.startswith("All "):
            continue
        worksheet = workbook[sheet_name]
        header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
        headers = [clean_text(value) for value in header_row]
        required = {"Word", "POS", "Nmorph", "PRS_signature", "MorphoLexSegm"}
        if not required.issubset(set(headers)):
            continue

        indexes = {name: headers.index(name) for name in required}
        for values in worksheet.iter_rows(min_row=2, values_only=True):
            word = clean_text(values[indexes["Word"]])
            segmentation = clean_text(values[indexes["MorphoLexSegm"]])
            if not word or not segmentation:
                continue
            prefixes, roots, suffixes = parse_parts(segmentation)
            key = lookup_key(word)
            entries.setdefault(
                key,
                MorphEntry(
                    word=word,
                    pos=clean_text(values[indexes["POS"]]),
                    nmorph=clean_text(values[indexes["Nmorph"]]),
                    prs_signature=clean_text(values[indexes["PRS_signature"]]),
                    segmentation=segmentation,
                    sheet=sheet_name,
                    prefixes=prefixes,
                    roots=roots,
                    suffixes=suffixes,
                ),
            )
    return entries


def update_source_register() -> None:
    fieldnames, rows = read_csv(SOURCES_PATH)
    source_row = {
        "source_key": MORPHOLEX_SOURCE_KEY,
        "source_category": "open_licensed",
        "source_name": MORPHOLEX_SOURCE_NAME,
        "source_url": MORPHOLEX_SOURCE_URL,
        "source_licence": MORPHOLEX_SOURCE_LICENCE,
        "source_use_note": MORPHOLEX_SOURCE_NOTE,
        "importability_status": "importable",
        "legal_review_status": "passed",
    }
    rows = [row for row in rows if row.get("source_key") != MORPHOLEX_SOURCE_KEY]
    rows.append(source_row)
    write_csv(SOURCES_PATH, fieldnames, rows)


def find_entry(entries: dict[str, MorphEntry], value: str) -> tuple[str, MorphEntry | None]:
    key = lookup_key(value)
    if key in entries:
        return "exact", entries[key]
    collapsed = collapsed_key(value)
    for entry_key, entry in entries.items():
        if collapsed_key(entry_key) == collapsed and collapsed:
            return "normalized_form", entry
    return "none", None


SPECIAL_COMPONENTS = {
    "ice_cream_en_gb": ("compound_components", ["ice", "cream"]),
    "it_s_en_gb": ("contraction_components", ["it", "is"]),
    "nature_natural_en_gb": ("combined_review_components", ["nature", "natural"]),
    "tall_taller_tallest_en_gb": ("combined_review_components", ["tall", "taller", "tallest"]),
    "they_re_en_gb": ("contraction_components", ["they", "are"]),
    "twenty_one_en_gb": ("compound_components", ["twenty", "one"]),
    "who_s_en_gb": ("contraction_components", ["who", "is"]),
    "you_re_en_gb": ("contraction_components", ["you", "are"]),
}


def select_evidence(word_row: dict[str, str], entries: dict[str, MorphEntry]) -> dict[str, Any]:
    for candidate in [word_row["normalised_word"], word_row["display_word"]]:
        match_type, entry = find_entry(entries, candidate)
        if entry:
            return {
                "match_type": match_type,
                "matched_forms": [entry.word],
                "entries": [entry],
                "selection_reason": (
                    "Exact MorphoLex row used."
                    if match_type == "exact"
                    else "Canonical spelling normalized to a MorphoLex row."
                ),
                "confidence": "medium" if match_type == "exact" else "low",
            }

    if word_row["word_key"] in SPECIAL_COMPONENTS:
        match_type, components = SPECIAL_COMPONENTS[word_row["word_key"]]
        component_entries = [entries.get(lookup_key(component)) for component in components]
        if all(component_entries):
            return {
                "match_type": match_type,
                "matched_forms": components,
                "entries": component_entries,
                "selection_reason": (
                    "No exact MorphoLex row; component MorphoLex rows used for "
                    "compound, contraction, or combined canonical review form."
                ),
                "confidence": "low",
            }

    return {
        "match_type": "unmatched",
        "matched_forms": [],
        "entries": [],
        "selection_reason": "No exact, normalized, or component MorphoLex evidence found.",
        "confidence": "",
    }


def morphemes_value(evidence: dict[str, Any]) -> str:
    entries: list[MorphEntry] = evidence["entries"]
    if not entries:
        return ""
    chunks = []
    for entry in entries:
        chunks.append(f"{entry.word}: " + " + ".join(ordered_parts(entry)))
    return "; ".join(chunks)


def morpholex_note(evidence: dict[str, Any]) -> str:
    entries: list[MorphEntry] = evidence["entries"]
    if not entries:
        return ""
    entry_notes = []
    for entry in entries:
        bits = [
            f"{entry.word}",
            f"POS={entry.pos or 'unlisted'}",
            f"Nmorph={entry.nmorph or 'unlisted'}",
            f"PRS={entry.prs_signature or 'unlisted'}",
            f"segmentation={entry.segmentation}",
        ]
        entry_notes.append("[" + "; ".join(bits) + "]")
    return (
        f"MorphoLex-en {evidence['match_type']} evidence: "
        + " ".join(entry_notes)
        + f" Selection note: {evidence['selection_reason']}"
    )


def merge_morphology_note(existing_note: str, new_note: str) -> str:
    if existing_note and new_note and new_note not in existing_note:
        return f"{existing_note} | {new_note}"
    return existing_note or new_note


def update_build_summary(summary: dict[str, Any]) -> None:
    if not BUILD_SUMMARY_PATH.exists():
        return
    data = json.loads(BUILD_SUMMARY_PATH.read_text(encoding="utf-8"))
    data.update(
        {
            "morpholex_morphology_metadata_populated": True,
            "morpholex_morphology_metadata_rows": summary["rows"],
            "morpholex_morphology_exact_rows": summary["match_type_counts"].get("exact", 0),
            "morpholex_morphology_normalized_rows": summary["match_type_counts"].get(
                "normalized_form", 0
            ),
            "morpholex_morphology_component_rows": summary["component_match_rows"],
            "morpholex_morphology_unmatched_rows": summary["match_type_counts"].get(
                "unmatched", 0
            ),
            "morpholex_source_importability_status": "importable",
            "morpholex_source_legal_review_status": "passed",
            "morphemes_populated_rows": summary["filled_field_counts"]["morphemes"],
            "morphology_notes_populated_rows": summary["filled_field_counts"][
                "morphology_notes"
            ],
            "sources": summary["source_rows_after_update"],
        }
    )
    BUILD_SUMMARY_PATH.write_text(
        json.dumps(data, indent=2, sort_keys=True) + "\n", encoding="utf-8"
    )


def main() -> None:
    metadata_fields, metadata_rows = read_csv(METADATA_PATH)
    _word_fields, word_rows = read_csv(WORDS_PATH)

    if [row["word_key"] for row in metadata_rows] != [row["word_key"] for row in word_rows]:
        raise SystemExit("canonical_word_metadata.csv row order does not match canonical_words.csv")
    if len(metadata_rows) != 874 or len({row["word_key"] for row in metadata_rows}) != 874:
        raise SystemExit("Expected 874 unique canonical metadata rows")

    entries = load_morpholex_entries(MORPHOLEX_XLSX_PATH)
    if not BEFORE_PATH.exists():
        shutil.copyfile(METADATA_PATH, BEFORE_PATH)
    update_source_register()

    intake_rows: list[dict[str, str]] = []
    complex_rows: list[dict[str, str]] = []
    audit_rows: list[dict[str, str]] = []
    match_type_counts: Counter[str] = Counter()
    confidence_counts: Counter[str] = Counter()
    filled_field_counts: Counter[str] = Counter()
    existing_note_preserved = 0
    pronunciation_changed = 0
    preserved_changed = 0

    for metadata_row, word_row in zip(metadata_rows, word_rows):
        before = dict(metadata_row)
        evidence = select_evidence(word_row, entries)
        match_type = evidence["match_type"]
        match_type_counts[match_type] += 1

        after_morphemes = morphemes_value(evidence)
        after_note = merge_morphology_note(
            metadata_row.get("morphology_notes", ""),
            morpholex_note(evidence),
        )
        if before.get("morphology_notes") and after_note != before.get("morphology_notes"):
            existing_note_preserved += 1

        if after_morphemes or after_note:
            metadata_row["morphemes"] = after_morphemes or metadata_row.get("morphemes", "")
            metadata_row["morphology_notes"] = after_note
            metadata_row["source_category"] = "open_licensed"
            metadata_row["source_name"] = MORPHOLEX_SOURCE_NAME
            metadata_row["source_url"] = MORPHOLEX_SOURCE_URL
            metadata_row["source_licence"] = MORPHOLEX_SOURCE_LICENCE
            metadata_row["source_use_note"] = (
                MORPHOLEX_SOURCE_NOTE
                + " Pronunciation values remain as previously populated and audited."
            )
            metadata_row["confidence"] = evidence["confidence"]
            metadata_row["review_status"] = "approved_for_first_exposure"

        for field in UNCHANGED_PRONUNCIATION_FIELDS:
            if metadata_row[field] != before[field]:
                pronunciation_changed += 1
        for field in PRESERVED_FIELDS:
            if metadata_row[field] != before[field]:
                preserved_changed += 1

        for field in ["morphemes", "morphology_notes"]:
            if metadata_row.get(field):
                filled_field_counts[field] += 1
        confidence_counts[metadata_row.get("confidence", "")] += 1

        entries_used: list[MorphEntry] = evidence["entries"]
        intake = {
            "word_key": word_row["word_key"],
            "normalised_word": word_row["normalised_word"],
            "display_word": word_row["display_word"],
            "match_type": match_type,
            "matched_forms": " | ".join(evidence["matched_forms"]),
            "selection_reason": evidence["selection_reason"],
            "morpholex_segments": " | ".join(entry.segmentation for entry in entries_used),
            "morpholex_pos": " | ".join(entry.pos for entry in entries_used),
            "morpholex_nmorph": " | ".join(entry.nmorph for entry in entries_used),
            "morpholex_prs_signature": " | ".join(
                entry.prs_signature for entry in entries_used
            ),
            "derived_morphemes": metadata_row.get("morphemes", ""),
            "derived_morphology_notes": metadata_row.get("morphology_notes", ""),
            "confidence": metadata_row.get("confidence", ""),
            "review_status": metadata_row.get("review_status", ""),
            "needs_human_review_note": (
                "Review non-exact component/normalized morphology before treating as "
                "final teaching explanation."
                if match_type != "exact"
                else ""
            ),
        }
        intake_rows.append(intake)
        if match_type != "exact":
            complex_rows.append(intake)

        audit_rows.append(
            {
                "word_key": word_row["word_key"],
                "match_type": match_type,
                "matched_forms": intake["matched_forms"],
                "selection_reason": evidence["selection_reason"],
                "before_morphemes": before.get("morphemes", ""),
                "after_morphemes": metadata_row.get("morphemes", ""),
                "before_morphology_notes": before.get("morphology_notes", ""),
                "after_morphology_notes": metadata_row.get("morphology_notes", ""),
                "before_source_name": before.get("source_name", ""),
                "after_source_name": metadata_row.get("source_name", ""),
                "before_confidence": before.get("confidence", ""),
                "after_confidence": metadata_row.get("confidence", ""),
                "before_review_status": before.get("review_status", ""),
                "after_review_status": metadata_row.get("review_status", ""),
                "pronunciation_fields_changed": "FALSE"
                if all(metadata_row[field] == before[field] for field in UNCHANGED_PRONUNCIATION_FIELDS)
                else "TRUE",
                "grapheme_irregularity_fields_changed": "FALSE"
                if all(metadata_row[field] == before[field] for field in PRESERVED_FIELDS)
                else "TRUE",
            }
        )

    if pronunciation_changed:
        raise SystemExit(f"Unexpected pronunciation field changes: {pronunciation_changed}")
    if preserved_changed:
        raise SystemExit(f"Unexpected grapheme/irregularity field changes: {preserved_changed}")

    write_csv(METADATA_PATH, metadata_fields, metadata_rows)
    write_csv(INTAKE_PATH, list(intake_rows[0].keys()), intake_rows)
    write_csv(COMPLEX_PATH, list(intake_rows[0].keys()), complex_rows)
    write_csv(AUDIT_PATH, list(audit_rows[0].keys()), audit_rows)

    _source_fields, source_rows = read_csv(SOURCES_PATH)
    summary = {
        "generated_at": date.today().isoformat(),
        "active_metadata_file": str(METADATA_PATH.relative_to(Path.cwd())),
        "before_snapshot": str(BEFORE_PATH.relative_to(Path.cwd())),
        "intake_file": str(INTAKE_PATH.relative_to(Path.cwd())),
        "complex_rows_file": str(COMPLEX_PATH.relative_to(Path.cwd())),
        "audit_file": str(AUDIT_PATH.relative_to(Path.cwd())),
        "morpholex_workbook_path_used": str(MORPHOLEX_XLSX_PATH),
        "rows": len(metadata_rows),
        "unique_word_keys": len({row["word_key"] for row in metadata_rows}),
        "row_order_matches_canonical_words": True,
        "match_type_counts": dict(match_type_counts),
        "component_match_rows": sum(
            count
            for key, count in match_type_counts.items()
            if key in {"compound_components", "contraction_components", "combined_review_components"}
        ),
        "confidence_counts": dict(confidence_counts),
        "filled_field_counts": dict(filled_field_counts),
        "existing_morphology_notes_preserved_and_augmented": existing_note_preserved,
        "pronunciation_fields_changed_count": pronunciation_changed,
        "grapheme_irregularity_fields_changed_count": preserved_changed,
        "review_status_set": "approved_for_first_exposure",
        "morpholex_source_importability_status": "importable",
        "morpholex_source_legal_review_status": "passed",
        "source_rows_after_update": len(source_rows),
        "hard_boundaries": [
            "no Supabase import",
            "no database mutation",
            "no migrations",
            "no importer changes",
            "no runtime hooks",
            "no resolver changes",
            "no assignment changes",
            "no evidence/proficiency changes",
            "no Word Treasure changes",
        ],
    }
    SUMMARY_PATH.write_text(json.dumps(summary, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    update_build_summary(summary)
    print(json.dumps(summary, indent=2, sort_keys=True))


if __name__ == "__main__":
    main()

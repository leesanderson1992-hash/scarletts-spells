#!/usr/bin/env python3
"""Legacy pre-simplification builder for Phase 5 source-intake candidate CSVs.

This script is deliberately local-file only. It reads existing repo-owned
source artifacts and writes review-candidate CSV exports into this dated folder.
It does not import data, connect to Supabase, mutate schema, or change runtime
code.

Do not run this script against the active candidate folder after the Teaching
Dictionary simplification. It writes the retired `canonical_word_micro_skills`
shape and static word-selection fields. The active reviewed CSVs now use
`micro_skill_word_support.csv`.
"""

from __future__ import annotations

import csv
import json
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[6]
CANDIDATE_DIR = Path(__file__).resolve().parent
CSV_DIR = CANDIDATE_DIR / "csv"

WORD_MAP = ROOT / "docs/implementation/seed-data/canonical-spelling-word-map/canonical-spelling-word-map-v1.xlsx"
D4_MICRO_SKILLS = ROOT / "docs/implementation/seed-data/domain4-seed-expansion/micro-skills.json"
COMMON_MISSPELLINGS = ROOT / "docs/implementation/seed-data/common_misspellings_seed_v1.csv"
TEACHING_TEMPLATE = ROOT / "docs/implementation/seed-data/teaching-dictionary/teaching-dictionary-workbook-template.xlsx"

UK_NC_URL = "https://assets.publishing.service.gov.uk/media/5a7ccc06ed915d63cc65ce61/English_Appendix_1_-_Spelling.pdf"

HEADERS = {
    "canonical_words.csv": [
        "word_key",
        "normalised_word",
        "display_word",
        "dialect_code",
        "frequency_band",
        "age_band",
        "complexity_band",
        "source_category",
        "source_name",
        "source_url",
        "source_licence",
        "source_use_note",
        "confidence",
        "review_status",
        "row_status",
    ],
    "canonical_word_metadata.csv": [
        "word_key",
        "syllables",
        "phoneme_hint",
        "grapheme_notes",
        "stress_pattern",
        "has_schwa",
        "morphemes",
        "morphology_notes",
        "irregularity_notes",
        "source_category",
        "source_name",
        "source_url",
        "source_licence",
        "source_use_note",
        "confidence",
        "review_status",
    ],
    "canonical_word_micro_skills.csv": [
        "word_key",
        "micro_skill_key",
        "micro_skill_role",
        "difficulty_band",
        "evidence_weight",
        "display_order",
        "source_category",
        "source_name",
        "source_url",
        "source_licence",
        "source_use_note",
        "confidence",
        "review_status",
    ],
    "teaching_content_versions.csv": [
        "micro_skill_key",
        "content_version",
        "version_status",
        "is_active",
        "teaching_objective",
        "child_friendly_explanation",
        "rule_explanation",
        "memory_tip",
        "anchor_word_key",
        "ordered_example_word_keys",
        "contrast_word_keys",
        "common_misconceptions",
        "first_exposure_progression",
        "review_progression",
        "source_category",
        "source_name",
        "source_url",
        "source_licence",
        "source_use_note",
        "confidence",
        "supersedes_content_version",
        "final_readiness_review_status",
        "final_readiness_reviewed_by",
        "final_readiness_reviewed_at",
    ],
    "teaching_content_field_reviews.csv": [
        "micro_skill_key",
        "content_version",
        "field_key",
        "review_gate",
        "review_status",
        "reviewed_by",
        "reviewed_at",
        "review_notes",
    ],
    "teaching_content_sources.csv": [
        "source_key",
        "source_category",
        "source_name",
        "source_url",
        "source_licence",
        "source_use_note",
        "importability_status",
        "legal_review_status",
    ],
}

SOURCE_NOTE_HEADERS = [
    "source_key",
    "source_category",
    "source_name",
    "source_url_or_path",
    "importability_status",
    "source_use_note",
    "supports_word_identity",
    "supports_metadata",
    "supports_mapping",
    "supports_teaching_copy",
]

UNREPRESENTED_HEADERS = [
    "source_key",
    "source_row_id",
    "source_field",
    "source_value",
    "related_word_key",
    "related_micro_skill_key",
    "reason",
]


SOURCE_FIELDS = {
    "word_map": {
        "source_category": "internal_reviewed_seed",
        "source_name": "Canonical spelling word-map workbook v1",
        "source_url": "docs/implementation/seed-data/canonical-spelling-word-map/canonical-spelling-word-map-v1.xlsx",
        "source_licence": "",
        "source_use_note": "Repo-owned internal pilot word-map content metadata; locally QA-audited; not resolver, assignment, mastery, or final teaching truth.",
        "confidence": "medium",
        "review_status": "in_review",
    },
    "d4_seed": {
        "source_category": "internal_reviewed_seed",
        "source_name": "Domain 4 seed expansion artifact",
        "source_url": "docs/implementation/seed-data/domain4-seed-expansion/micro-skills.json",
        "source_licence": "",
        "source_use_note": "Repo-owned D4 seed artifact for existing micro-skill keys, starter word banks, and seed teaching points; candidate metadata only.",
        "confidence": "medium",
        "review_status": "in_review",
    },
    "common_misspellings": {
        "source_category": "internal_reviewed_seed",
        "source_name": "Common misspellings seed v1",
        "source_url": "docs/implementation/seed-data/common_misspellings_seed_v1.csv",
        "source_licence": "",
        "source_use_note": "Repo-owned curated misspelling examples; candidate evidence only, not learner evidence, resolver truth, or final teaching truth.",
        "confidence": "medium",
        "review_status": "in_review",
    },
    "ai_draft": {
        "source_category": "ai_assisted_draft",
        "source_name": "Phase 5 source-intake AI-assisted draft",
        "source_url": "docs/implementation/seed-data/teaching-dictionary/candidates/2026-06-29-phase-5-source-intake",
        "source_licence": "",
        "source_use_note": "AI-assisted draft copy for human curriculum review only; must not be promoted or surfaced without human approval.",
        "confidence": "low",
        "review_status": "ai_draft",
    },
}


def clean(value: Any) -> str:
    return "" if value is None else str(value).strip()


def word_key(word: str, dialect: str = "en_gb") -> str:
    base = re.sub(r"[^a-z0-9]+", "_", word.lower()).strip("_")
    return f"{base}_{dialect}"


def frequency_band(value: str) -> str:
    return {"common": "high", "medium": "medium", "rare": "low"}.get(value.lower(), "medium")


def complexity_band(value: str) -> str:
    return {"easy": "low", "medium": "medium", "hard": "high", "seed": "medium"}.get(value.lower(), "medium")


def role_from_word_map(row: dict[str, Any]) -> str:
    role = clean(row.get("word_role")).lower()
    if role in {"teaching_example", "anchor_word"}:
        return "anchor"
    if role == "review_word":
        return "review_example"
    return "ordered_example"


def write_csv(path: Path, headers: list[str], rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers)
        writer.writeheader()
        for row in rows:
            writer.writerow({header: clean(row.get(header)) for header in headers})


def worksheet_rows(path: Path, sheet_name: str) -> list[dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    values = list(ws.iter_rows(values_only=True))
    headers = [clean(value) for value in values[0]]
    rows = []
    for values_row in values[1:]:
        row = dict(zip(headers, values_row))
        if any(value is not None and clean(value) for value in values_row):
            rows.append(row)
    return rows


def source_register_rows() -> list[dict[str, Any]]:
    return [
        {
            "source_key": "canonical_spelling_word_map_v1",
            "source_category": "internal_reviewed_seed",
            "source_name": "Canonical spelling word-map workbook v1",
            "source_url": "docs/implementation/seed-data/canonical-spelling-word-map/canonical-spelling-word-map-v1.xlsx",
            "source_licence": "internal/manual",
            "source_use_note": "Internal pilot content metadata; locally QA-audited; not runtime, resolver, mastery, assignment, or final teaching truth.",
            "importability_status": "importable",
            "legal_review_status": "not_required",
        },
        {
            "source_key": "domain4_seed_expansion",
            "source_category": "internal_reviewed_seed",
            "source_name": "Domain 4 seed expansion artifact",
            "source_url": "docs/implementation/seed-data/domain4-seed-expansion/micro-skills.json",
            "source_licence": "internal",
            "source_use_note": "Internal seed artifact for existing D4 keys, starter words, and seed teaching points; candidate metadata only.",
            "importability_status": "importable",
            "legal_review_status": "not_required",
        },
        {
            "source_key": "common_misspellings_seed_v1",
            "source_category": "internal_reviewed_seed",
            "source_name": "Common misspellings seed v1",
            "source_url": "docs/implementation/seed-data/common_misspellings_seed_v1.csv",
            "source_licence": "internal",
            "source_use_note": "Internal curated examples for candidate word identity and diagnostic mapping notes; not learner evidence or resolver truth.",
            "importability_status": "importable",
            "legal_review_status": "not_required",
        },
        {
            "source_key": "teaching_dictionary_workbook_template",
            "source_category": "internal_authored",
            "source_name": "Phase 5C teaching dictionary workbook template",
            "source_url": "docs/implementation/seed-data/teaching-dictionary/teaching-dictionary-workbook-template.xlsx",
            "source_licence": "internal",
            "source_use_note": "Internal template and CSV contract helper; schema contract only, not teaching content truth.",
            "importability_status": "importable",
            "legal_review_status": "not_required",
        },
        {
            "source_key": "phase_5_validator_fixtures",
            "source_category": "internal_authored",
            "source_name": "Phase 5C validator fixtures",
            "source_url": "scripts/fixtures/teaching-dictionary-csv",
            "source_licence": "internal",
            "source_use_note": "Synthetic validator-contract examples inspected for shape only; not imported as real teaching content.",
            "importability_status": "not_importable",
            "legal_review_status": "not_required",
        },
        {
            "source_key": "uk_national_curriculum_spelling_appendix",
            "source_category": "reference_only",
            "source_name": "UK National Curriculum English spelling appendix",
            "source_url": UK_NC_URL,
            "source_licence": "",
            "source_use_note": "Curriculum-scope reference only unless importability is separately established; not surfaced as child-facing copy.",
            "importability_status": "reference_only",
            "legal_review_status": "unknown",
        },
        {
            "source_key": "phase_5_source_intake_ai_draft",
            "source_category": "ai_assisted_draft",
            "source_name": "Phase 5 source-intake AI-assisted draft",
            "source_url": "docs/implementation/seed-data/teaching-dictionary/candidates/2026-06-29-phase-5-source-intake",
            "source_licence": "",
            "source_use_note": "AI-assisted draft explanation/progression text for human review only; not approved or final teaching truth.",
            "importability_status": "unknown",
            "legal_review_status": "required",
        },
    ]


def source_note_rows() -> list[dict[str, Any]]:
    supports = {
        "canonical_spelling_word_map_v1": ("TRUE", "TRUE", "TRUE", "FALSE"),
        "domain4_seed_expansion": ("TRUE", "FALSE", "TRUE", "TRUE"),
        "common_misspellings_seed_v1": ("TRUE", "FALSE", "TRUE", "FALSE"),
        "teaching_dictionary_workbook_template": ("FALSE", "FALSE", "FALSE", "FALSE"),
        "phase_5_validator_fixtures": ("FALSE", "FALSE", "FALSE", "FALSE"),
        "uk_national_curriculum_spelling_appendix": ("FALSE", "FALSE", "FALSE", "FALSE"),
        "phase_5_source_intake_ai_draft": ("FALSE", "FALSE", "FALSE", "TRUE"),
    }
    rows = []
    for row in source_register_rows():
        word_identity, metadata, mapping, teaching_copy = supports[row["source_key"]]
        rows.append(
            {
                "source_key": row["source_key"],
                "source_category": row["source_category"],
                "source_name": row["source_name"],
                "source_url_or_path": row["source_url"],
                "importability_status": row["importability_status"],
                "source_use_note": row["source_use_note"],
                "supports_word_identity": word_identity,
                "supports_metadata": metadata,
                "supports_mapping": mapping,
                "supports_teaching_copy": teaching_copy,
            }
        )
    return rows


def main() -> None:
    sys.exit(
        "This legacy builder writes the retired pre-simplification CSV shape. "
        "Use the existing simplified csv/ folder instead."
    )

    CSV_DIR.mkdir(parents=True, exist_ok=True)

    d4_rows = json.loads(D4_MICRO_SKILLS.read_text(encoding="utf-8"))
    d4_by_key = {row["micro_skill_key"]: row for row in d4_rows}
    word_bank = worksheet_rows(WORD_MAP, "micro_skill_word_bank")
    word_metadata = worksheet_rows(WORD_MAP, "word_metadata")
    contrast_pairs = worksheet_rows(WORD_MAP, "contrast_pairs")
    diagnostic_mappings = worksheet_rows(WORD_MAP, "diagnostic_misspelling_mappings")
    common_rows = list(csv.DictReader(COMMON_MISSPELLINGS.open("r", encoding="utf-8")))

    words: dict[str, dict[str, Any]] = {}
    metadata: dict[str, dict[str, Any]] = {}
    mappings: dict[tuple[str, str, str, str], dict[str, Any]] = {}
    unrepresented: list[dict[str, Any]] = []

    def ensure_word(
        word: str,
        source: str,
        *,
        frequency: str = "medium",
        complexity: str = "medium",
        confidence: str | None = None,
        review_status: str | None = None,
    ) -> str:
        normalised = word.lower().strip()
        key = word_key(normalised)
        source_info = SOURCE_FIELDS[source]
        if key not in words:
            words[key] = {
                "word_key": key,
                "normalised_word": normalised,
                "display_word": normalised,
                "dialect_code": "en-GB",
                "frequency_band": frequency,
                "age_band": "candidate_review",
                "complexity_band": complexity,
                **source_info,
                "confidence": confidence or source_info["confidence"],
                "review_status": review_status or source_info["review_status"],
                "row_status": "draft",
            }
        return key

    def ensure_metadata(key: str, source: str, **values: Any) -> None:
        source_info = SOURCE_FIELDS[source]
        if key not in metadata:
            metadata[key] = {
                "word_key": key,
                "syllables": "",
                "phoneme_hint": "",
                "grapheme_notes": "",
                "stress_pattern": "",
                "has_schwa": "FALSE",
                "morphemes": "",
                "morphology_notes": "",
                "irregularity_notes": "",
                **source_info,
                "confidence": values.pop("confidence", source_info["confidence"]),
                "review_status": values.pop("review_status", source_info["review_status"]),
            }
        for field, value in values.items():
            if clean(value):
                metadata[key][field] = value

    def add_mapping(
        key: str,
        micro_skill_key: str,
        role: str,
        source: str,
        *,
        order: int,
        difficulty: str = "medium",
        evidence_weight: str = "1",
        confidence: str | None = None,
        review_status: str | None = None,
        note_suffix: str = "",
    ) -> None:
        source_info = dict(SOURCE_FIELDS[source])
        if note_suffix:
            source_info["source_use_note"] = f"{source_info['source_use_note']} {note_suffix}"
        map_key = (key, micro_skill_key, role, source_info["source_name"])
        existing = mappings.get(map_key)
        if existing is not None:
            existing["source_use_note"] = source_info["source_use_note"]
            return
        mappings[map_key] = {
            "word_key": key,
            "micro_skill_key": micro_skill_key,
            "micro_skill_role": role,
            "difficulty_band": difficulty,
            "evidence_weight": evidence_weight,
            "display_order": order,
            **source_info,
            "confidence": confidence or source_info["confidence"],
            "review_status": review_status or source_info["review_status"],
        }

    for row in word_metadata:
        key = ensure_word(
            clean(row["normalised_word"]),
            "word_map",
            complexity=complexity_band(clean(row.get("irregularity_band"))),
        )
        ensure_metadata(
            key,
            "word_map",
            syllables=clean(row.get("syllable_count")),
            phoneme_hint=clean(row.get("phoneme_hint")),
            stress_pattern=clean(row.get("stress_pattern")),
            has_schwa="TRUE" if row.get("has_schwa") is True else "FALSE",
            morphology_notes=clean(row.get("morphology_notes")),
            irregularity_notes=clean(row.get("irregularity_band")),
        )

    per_skill_order = Counter()
    for row in word_bank:
        skill = clean(row["micro_skill_key"])
        key = ensure_word(
            clean(row["normalised_word"]),
            "word_map",
            frequency=frequency_band(clean(row.get("frequency_band"))),
            complexity=complexity_band(clean(row.get("complexity_band"))),
        )
        ensure_metadata(key, "word_map")
        per_skill_order[skill] += 1
        add_mapping(
            key,
            skill,
            role_from_word_map(row),
            "word_map",
            order=per_skill_order[skill],
            difficulty=complexity_band(clean(row.get("complexity_band"))),
            evidence_weight="1",
        )

    for index, row in enumerate(contrast_pairs, start=1):
        target_skill = clean(row["target_micro_skill_key"])
        target_key = ensure_word(clean(row["target_word"]), "word_map")
        contrast_key = ensure_word(clean(row["contrast_word"]), "word_map")
        ensure_metadata(target_key, "word_map")
        ensure_metadata(contrast_key, "word_map")
        add_mapping(
            contrast_key,
            target_skill,
            "contrast",
            "word_map",
            order=1000 + index,
            difficulty="medium",
            evidence_weight="0.5",
            note_suffix=f"Contrast pair candidate for {clean(row['target_word'])} vs {clean(row['contrast_word'])}; not final teaching truth.",
        )

    for index, row in enumerate(diagnostic_mappings, start=1):
        skill = clean(row["micro_skill_key"])
        correction = clean(row["correction_normalised"])
        key = ensure_word(correction, "word_map")
        ensure_metadata(key, "word_map")
        add_mapping(
            key,
            skill,
            "diagnostic",
            "word_map",
            order=2000 + index,
            difficulty="medium",
            evidence_weight="0.5",
            confidence=clean(row.get("confidence")) or "medium",
            note_suffix=f"Diagnostic misspelling {clean(row['misspelling_normalised'])} -> {correction}; resolver_visible_candidate remains FALSE.",
        )
        unrepresented.append(
            {
                "source_key": "canonical_spelling_word_map_v1",
                "source_row_id": f"diagnostic_misspelling_mappings:{index}",
                "source_field": "misspelling_normalised",
                "source_value": clean(row["misspelling_normalised"]),
                "related_word_key": key,
                "related_micro_skill_key": skill,
                "reason": "Phase 5C has no canonical_misspellings.csv sheet; misspelling is preserved only in mapping source notes for review.",
            }
        )

    for skill_index, skill_row in enumerate(d4_rows, start=1):
        skill = skill_row["micro_skill_key"]
        meta = skill_row.get("metadata", {})
        starter_rows = meta.get("starter_word_bank") or []
        example_words = meta.get("example_words") or []
        source_words = [clean(row.get("word")) for row in starter_rows if clean(row.get("word"))] or [
            clean(word) for word in example_words if clean(word)
        ]
        for order, word in enumerate(source_words, start=1):
            key = ensure_word(word, "d4_seed", complexity="medium")
            ensure_metadata(key, "d4_seed")
            add_mapping(
                key,
                skill,
                "anchor" if order == 1 else "ordered_example",
                "d4_seed",
                order=order,
                difficulty="medium",
                evidence_weight="1" if order == 1 else "0.75",
            )

    common_grouped: dict[tuple[str, str], list[dict[str, str]]] = defaultdict(list)
    for row in common_rows:
        common_grouped[(row["correction"].strip().lower(), row["suggested_micro_skill_key"].strip())].append(row)

    for index, ((correction, skill), rows) in enumerate(sorted(common_grouped.items()), start=1):
        key = ensure_word(correction, "common_misspellings", confidence="medium", review_status="in_review")
        ensure_metadata(
            key,
            "common_misspellings",
            irregularity_notes=rows[0].get("pattern_hint", ""),
            confidence="medium",
            review_status="in_review",
        )
        sample_misspellings = ", ".join(row["misspelling"] for row in rows[:4])
        add_mapping(
            key,
            skill,
            "diagnostic",
            "common_misspellings",
            order=3000 + index,
            difficulty="medium",
            evidence_weight="0.5",
            confidence="medium",
            review_status="in_review",
            note_suffix=f"Aggregates {len(rows)} misspelling seed row(s), including: {sample_misspellings}.",
        )
        for row in rows:
            unrepresented.append(
                {
                    "source_key": "common_misspellings_seed_v1",
                    "source_row_id": row.get("source_row_id", ""),
                    "source_field": "misspelling",
                    "source_value": row.get("misspelling", ""),
                    "related_word_key": key,
                    "related_micro_skill_key": skill,
                    "reason": "Phase 5C has no canonical_misspellings.csv sheet; misspelling is represented only as source-note evidence on a diagnostic candidate mapping.",
                }
            )

    teaching_versions: list[dict[str, Any]] = []
    field_reviews: list[dict[str, Any]] = []
    mappings_by_skill_role: dict[tuple[str, str], list[str]] = defaultdict(list)
    for mapping in sorted(mappings.values(), key=lambda row: (row["micro_skill_key"], int(row["display_order"]), row["word_key"])):
        mappings_by_skill_role[(mapping["micro_skill_key"], mapping["micro_skill_role"])].append(mapping["word_key"])

    for skill_row in d4_rows:
        skill = skill_row["micro_skill_key"]
        meta = skill_row.get("metadata", {})
        anchor_words = mappings_by_skill_role.get((skill, "anchor"), [])
        ordered_words = mappings_by_skill_role.get((skill, "ordered_example"), [])
        contrast_words = mappings_by_skill_role.get((skill, "contrast"), [])
        anchor_key = anchor_words[0] if anchor_words else (ordered_words[0] if ordered_words else "")
        ordered_keys = []
        for key in anchor_words + ordered_words:
            if key not in ordered_keys:
                ordered_keys.append(key)
        teaching_point = clean(meta.get("teaching_point"))
        display_name = clean(skill_row.get("display_name"))
        teaching_versions.append(
            {
                "micro_skill_key": skill,
                "content_version": "draft_source_intake_v1",
                "version_status": "draft",
                "is_active": "FALSE",
                "teaching_objective": f"Draft objective: {display_name}." if display_name else f"Draft objective for {skill}.",
                "child_friendly_explanation": f"Draft for review: notice the spelling pattern in {words[anchor_key]['display_word']} and compare it with the examples."
                if anchor_key
                else "",
                "rule_explanation": teaching_point or f"Draft seed rule for {display_name}.",
                "memory_tip": "",
                "anchor_word_key": anchor_key,
                "ordered_example_word_keys": "|".join(ordered_keys[:8]),
                "contrast_word_keys": "|".join(contrast_words[:6]),
                "common_misconceptions": "Draft misconception notes should be added by human review.",
                "first_exposure_progression": "rule_explanation|guided_rule_application",
                "review_progression": "rapid_recall",
                **SOURCE_FIELDS["ai_draft"],
                "supersedes_content_version": "",
                "final_readiness_review_status": "not_started",
                "final_readiness_reviewed_by": "",
                "final_readiness_reviewed_at": "",
            }
        )
        for field_key, gate in [
            ("teaching_objective", "pedagogy"),
            ("child_friendly_explanation", "child_language"),
            ("rule_explanation", "pedagogy"),
            ("anchor_word_key", "pedagogy"),
            ("ordered_example_word_keys", "pedagogy"),
            ("first_exposure_progression", "pedagogy"),
            ("review_progression", "pedagogy"),
            ("source", "source_licence"),
            ("licence", "source_licence"),
        ]:
            field_reviews.append(
                {
                    "micro_skill_key": skill,
                    "content_version": "draft_source_intake_v1",
                    "field_key": field_key,
                    "review_gate": gate,
                    "review_status": "ai_draft" if field_key not in {"anchor_word_key", "ordered_example_word_keys"} else "in_review",
                    "reviewed_by": "",
                    "reviewed_at": "",
                    "review_notes": "Candidate source-intake row only; requires human review before any approval or surfacing.",
                }
            )

    word_rows = sorted(words.values(), key=lambda row: row["word_key"])
    metadata_rows = [metadata.get(row["word_key"]) or {"word_key": row["word_key"], **SOURCE_FIELDS["d4_seed"]} for row in word_rows]
    mapping_rows = sorted(mappings.values(), key=lambda row: (row["micro_skill_key"], int(row["display_order"]), row["word_key"], row["micro_skill_role"]))

    write_csv(CSV_DIR / "teaching_content_sources.csv", HEADERS["teaching_content_sources.csv"], source_register_rows())
    write_csv(CSV_DIR / "canonical_words.csv", HEADERS["canonical_words.csv"], word_rows)
    write_csv(CSV_DIR / "canonical_word_metadata.csv", HEADERS["canonical_word_metadata.csv"], metadata_rows)
    write_csv(CSV_DIR / "canonical_word_micro_skills.csv", HEADERS["canonical_word_micro_skills.csv"], mapping_rows)
    write_csv(CSV_DIR / "teaching_content_versions.csv", HEADERS["teaching_content_versions.csv"], teaching_versions)
    write_csv(CSV_DIR / "teaching_content_field_reviews.csv", HEADERS["teaching_content_field_reviews.csv"], field_reviews)

    write_csv(CANDIDATE_DIR / "source_register_notes.csv", SOURCE_NOTE_HEADERS, source_note_rows())
    write_csv(CANDIDATE_DIR / "unrepresented_rows.csv", UNREPRESENTED_HEADERS, unrepresented)

    summary = {
        "words": len(word_rows),
        "metadata_rows": len(metadata_rows),
        "mapping_rows": len(mapping_rows),
        "teaching_content_versions": len(teaching_versions),
        "field_review_rows": len(field_reviews),
        "sources": len(source_register_rows()),
        "unrepresented_rows": len(unrepresented),
        "micro_skills_represented_in_mappings": len({row["micro_skill_key"] for row in mapping_rows}),
        "micro_skills_with_teaching_versions": len(teaching_versions),
        "source_categories": sorted({row["source_category"] for row in source_register_rows()}),
        "importability_statuses": sorted({row["importability_status"] for row in source_register_rows()}),
    }
    (CANDIDATE_DIR / "build_summary.json").write_text(json.dumps(summary, indent=2) + "\n", encoding="utf-8")

    readme = f"""# Phase 5 Source-Intake Candidate

Generated from local repo-owned seed artifacts on 2026-06-29.

This candidate folder is draft data only. It does not approve teaching content,
import into Supabase, mutate schema, create runtime hooks, change resolver
behavior, write evidence/proficiency, affect assignments, or change Word
Treasure.

## Contents

- `csv/`: Phase 5C validator CSV export folder.
- `source_register_notes.csv`: richer source register including support flags
  that do not fit the Phase 5C `teaching_content_sources.csv` header.
- `unrepresented_rows.csv`: misspelling/example source fields that cannot be
  represented directly by the current Phase 5C CSV contract.
- `build_summary.json`: generated row counts.

## Generated Counts

- canonical words: {summary["words"]}
- metadata rows: {summary["metadata_rows"]}
- word-to-micro-skill mappings: {summary["mapping_rows"]}
- teaching content versions: {summary["teaching_content_versions"]}
- field review rows: {summary["field_review_rows"]}
- source rows: {summary["sources"]}
- rows/fields not directly representable: {summary["unrepresented_rows"]}

All review/content statuses remain `draft`, `in_review`, or `ai_draft`.
No `approved_for_first_exposure`, `approved_for_guided_review`, or `signed_off`
status is assigned to real candidate content.
"""
    (CANDIDATE_DIR / "README.md").write_text(readme, encoding="utf-8")

    print(json.dumps(summary, indent=2))


if __name__ == "__main__":
    main()

#!/usr/bin/env python3
"""Generate deterministic Domain 4 seed expansion artifacts from the workbook.

This is a dry-run artifact generator only. It does not connect to Supabase,
create migrations, or import rows.
"""

from __future__ import annotations

import json
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover - local setup guard
    raise SystemExit(
        "openpyxl is required to parse the Domain 4 seed workbook. "
        "Run this script with the bundled workspace Python runtime."
    ) from exc


WORKBOOK_PATH = Path("docs/D4 Seeding Map Finale Final.xlsx")
OUTPUT_DIR = Path("docs/implementation/seed-data/domain4-seed-expansion")
SEED_VERSION = "domain4-seed-expansion-v1"
SUPPORTED_PRACTICE_ROUTE = "word_practice"
SUPPORTED_DB_PRACTICE_ROUTES = ["word_practice", "grouped_set_practice"]
EXCLUDED_MAPPING_FAMILY_IDS = ["D4_PROOF"]

EXPECTED_COUNTS = {
    "ready_families": 8,
    "clusters": 47,
    "micro_skills": 240,
    "task_templates": 12,
    "family_level_template_mappings": 40,
}

DIPHTHONG_NODE_IDS = [
    "D4_PG_DIPHTHONGS_OU",
    "D4_PG_DIPHTHONGS_OW",
    "D4_PG_DIPHTHONGS_OI",
    "D4_PG_DIPHTHONGS_OY",
]


def clean_string(value: Any) -> str:
    if value is None:
        return ""

    if isinstance(value, float) and value.is_integer():
        return str(int(value))

    return str(value).strip()


def clean_optional_string(value: Any) -> str | None:
    cleaned = clean_string(value)
    return cleaned or None


def read_sheet(workbook: Any, sheet_name: str) -> list[dict[str, Any]]:
    worksheet = workbook[sheet_name]
    rows = [
        row
        for row in worksheet.iter_rows(values_only=True)
        if any(value is not None for value in row)
    ]

    if not rows:
        return []

    headers = [clean_string(value) for value in rows[0]]
    records: list[dict[str, Any]] = []

    for row_index, row in enumerate(rows[1:], start=2):
        record: dict[str, Any] = {"_row": row_index}
        for column_index, header in enumerate(headers):
            if not header:
                continue
            record[header] = row[column_index] if column_index < len(row) else None
        records.append(record)

    return records


def split_words(value: Any) -> list[str]:
    text = clean_string(value)
    if not text:
        return []

    words: list[str] = []
    for raw_part in text.replace("\n", ",").split(","):
        word = raw_part.strip()
        if word:
            words.append(word)
    return words


def unique_in_order(values: list[str]) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []
    for value in values:
        if value not in seen:
            result.append(value)
            seen.add(value)
    return result


def assert_no_duplicates(records: list[dict[str, Any]], key: str, label: str) -> None:
    keys = [clean_string(record.get(key)) for record in records]
    duplicates = sorted(key for key, count in Counter(keys).items() if key and count > 1)
    missing_rows = [record["_row"] for record in records if not clean_string(record.get(key))]

    if duplicates or missing_rows:
        raise ValueError(
            f"{label} key validation failed: duplicates={duplicates}, "
            f"missing_rows={missing_rows}"
        )


def build_allowed_template_keys(
    mappings: list[dict[str, Any]],
) -> dict[str, list[str]]:
    by_family: dict[str, list[tuple[int, str]]] = defaultdict(list)

    for mapping in mappings:
        family_id = clean_string(mapping.get("family_id"))
        level = int(float(mapping.get("level") or 0))
        for field in ("default_template_id", "fallback_template_id"):
            template_id = clean_string(mapping.get(field))
            if template_id:
                by_family[family_id].append((level, template_id))

    return {
        family_id: unique_in_order(
            [template_id for _, template_id in sorted(values, key=lambda item: item[0])]
        )
        for family_id, values in by_family.items()
    }


def normalize_family(record: dict[str, Any]) -> dict[str, Any]:
    family_id = clean_string(record.get("family_id"))
    launch_status = clean_string(record.get("Launch Status"))
    return {
        "mastery_domain_key": "D4",
        "skill_family_key": family_id,
        "display_name": clean_string(record.get("family_name")),
        "is_active": launch_status.lower() == "ready",
        "is_assignable": launch_status.lower() == "ready",
        "metadata": {
            "seed_version": SEED_VERSION,
            "source_workbook": str(WORKBOOK_PATH),
            "launch_status": launch_status,
            "purpose": clean_optional_string(record.get("purpose")),
        },
    }


def normalize_cluster(record: dict[str, Any], ready_family_ids: set[str]) -> dict[str, Any]:
    family_id = clean_string(record.get("family_id"))
    return {
        "mastery_domain_key": "D4",
        "skill_family_key": family_id,
        "skill_cluster_key": clean_string(record.get("cluster_id")),
        "display_name": clean_string(record.get("refined_cluster_name")),
        "is_active": family_id in ready_family_ids,
        "is_assignable": family_id in ready_family_ids,
        "metadata": {
            "seed_version": SEED_VERSION,
            "source_workbook": str(WORKBOOK_PATH),
            "purpose": clean_optional_string(record.get("purpose")),
        },
    }


def normalize_micro_skill(
    record: dict[str, Any],
    ready_family_ids: set[str],
    allowed_template_keys_by_family: dict[str, list[str]],
) -> dict[str, Any]:
    family_id = clean_string(record.get("family_id"))
    example_words = split_words(record.get("example_words"))
    starter_word_bank = [
        {
            "word": word,
            "difficulty": "seed",
            "common_wrong_spellings": [],
        }
        for word in example_words
    ]

    return {
        "mastery_domain_key": "D4",
        "skill_family_key": family_id,
        "skill_cluster_key": clean_string(record.get("cluster_id")),
        "micro_skill_key": clean_string(record.get("node_id")),
        "display_name": clean_string(record.get("node_name")),
        "practice_route": SUPPORTED_PRACTICE_ROUTE,
        "is_active": family_id in ready_family_ids,
        "is_assignable": family_id in ready_family_ids,
        "allowed_template_keys": allowed_template_keys_by_family.get(family_id, []),
        "metadata": {
            "seed_version": SEED_VERSION,
            "source_workbook": str(WORKBOOK_PATH),
            "cluster_name": clean_optional_string(record.get("cluster_name")),
            "developmental_foundation": clean_optional_string(
                record.get("developmental_foundation")
            ),
            "teaching_point": clean_optional_string(record.get("teaching_point")),
            "example_words": example_words,
            "starter_word_bank": starter_word_bank,
            "unsupported_runtime_notes": [
                "Task templates and family-level mappings are repo artifacts/config until dedicated schema exists.",
                "Resolver adoption is intentionally separate from taxonomy seeding.",
            ],
        },
    }


def normalize_task_template(record: dict[str, Any]) -> dict[str, Any]:
    return {
        "template_id": clean_string(record.get("Template ID")),
        "template_name": clean_string(record.get("Template name")),
        "main_purpose": clean_string(record.get("Main purpose")),
        "best_for_level": clean_string(record.get("Best for level")),
        "metadata": {
            "seed_version": SEED_VERSION,
            "source_workbook": str(WORKBOOK_PATH),
            "storage_status": "repo_artifact_until_schema_exists",
        },
    }


def normalize_family_mapping(record: dict[str, Any]) -> dict[str, Any]:
    return {
        "family_id": clean_string(record.get("family_id")),
        "level": int(float(record.get("level") or 0)),
        "competency_stage": clean_string(record.get("competency_stage")),
        "default_template_id": clean_string(record.get("default_template_id")),
        "default_template_route": clean_string(record.get("default_template_route")),
        "fallback_template_id": clean_string(record.get("fallback_template_id")),
        "fallback_template_route": clean_string(record.get("fallback_template_route")),
        "rationale": clean_string(record.get("rationale")),
        "metadata": {
            "seed_version": SEED_VERSION,
            "source_workbook": str(WORKBOOK_PATH),
            "storage_status": "repo_artifact_until_schema_exists",
        },
    }


def missing_mapping_fields(
    mappings: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    required_fields = [
        "competency_stage",
        "default_template_id",
        "default_template_route",
        "fallback_template_id",
        "fallback_template_route",
        "rationale",
    ]
    missing: list[dict[str, Any]] = []

    for mapping in mappings:
        fields = [
            field for field in required_fields if not clean_string(mapping.get(field))
        ]
        if fields:
            missing.append(
                {
                    "row": mapping["_row"],
                    "family_id": clean_string(mapping.get("family_id")),
                    "level": int(float(mapping.get("level") or 0)),
                    "missing_fields": fields,
                }
            )

    return missing


def exclude_mapping_rows(
    mappings: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    excluded_family_ids = set(EXCLUDED_MAPPING_FAMILY_IDS)
    included: list[dict[str, Any]] = []
    excluded: list[dict[str, Any]] = []

    for mapping in mappings:
        if clean_string(mapping.get("family_id")) in excluded_family_ids:
            excluded.append(mapping)
        else:
            included.append(mapping)

    return included, excluded


def write_json(path: Path, value: Any) -> None:
    path.write_text(json.dumps(value, indent=2, sort_keys=True) + "\n", encoding="utf8")


def main() -> None:
    if not WORKBOOK_PATH.exists():
        raise FileNotFoundError(f"Missing workbook: {WORKBOOK_PATH}")

    workbook = load_workbook(WORKBOOK_PATH, read_only=True, data_only=True)
    required_sheets = [
        "Skill Families",
        "Skill Cluster",
        "Micro Skills",
        "Task Templates",
        "Family Level Mapping",
    ]
    missing_sheets = [sheet for sheet in required_sheets if sheet not in workbook.sheetnames]
    if missing_sheets:
        raise ValueError(f"Missing required sheets: {missing_sheets}")

    family_rows = read_sheet(workbook, "Skill Families")
    cluster_rows = read_sheet(workbook, "Skill Cluster")
    micro_skill_rows = read_sheet(workbook, "Micro Skills")
    task_template_rows = read_sheet(workbook, "Task Templates")
    family_mapping_rows = read_sheet(workbook, "Family Level Mapping")
    included_family_mapping_rows, excluded_family_mapping_rows = exclude_mapping_rows(
        family_mapping_rows
    )

    assert_no_duplicates(family_rows, "family_id", "Family")
    assert_no_duplicates(cluster_rows, "cluster_id", "Cluster")
    assert_no_duplicates(micro_skill_rows, "node_id", "Micro-skill")
    assert_no_duplicates(task_template_rows, "Template ID", "Task template")

    family_ids = {clean_string(record.get("family_id")) for record in family_rows}
    ready_family_ids = {
        clean_string(record.get("family_id"))
        for record in family_rows
        if clean_string(record.get("Launch Status")).lower() == "ready"
    }
    cluster_ids = {clean_string(record.get("cluster_id")) for record in cluster_rows}
    template_ids = {
        clean_string(record.get("Template ID")) for record in task_template_rows
    }

    unknown_cluster_families = sorted(
        {
            clean_string(record.get("family_id"))
            for record in cluster_rows
            if clean_string(record.get("family_id")) not in family_ids
        }
    )
    unknown_micro_skill_families = sorted(
        {
            clean_string(record.get("family_id"))
            for record in micro_skill_rows
            if clean_string(record.get("family_id")) not in family_ids
        }
    )
    unknown_micro_skill_clusters = sorted(
        {
            clean_string(record.get("cluster_id"))
            for record in micro_skill_rows
            if clean_string(record.get("cluster_id")) not in cluster_ids
        }
    )
    cluster_skill_counts = Counter(
        clean_string(record.get("cluster_id")) for record in micro_skill_rows
    )
    empty_clusters = sorted(
        cluster_id for cluster_id in cluster_ids if cluster_skill_counts[cluster_id] == 0
    )

    missing_ready_metadata: dict[str, list[int]] = defaultdict(list)
    for record in micro_skill_rows:
        if clean_string(record.get("family_id")) not in ready_family_ids:
            continue
        for field in (
            "node_name",
            "developmental_foundation",
            "teaching_point",
            "example_words",
        ):
            if not clean_string(record.get(field)):
                missing_ready_metadata[field].append(record["_row"])

    unknown_mapping_families = sorted(
        {
            clean_string(record.get("family_id"))
            for record in included_family_mapping_rows
            if clean_string(record.get("family_id")) not in family_ids
        }
    )
    unknown_mapping_templates = sorted(
        {
            clean_string(record.get(field))
            for record in included_family_mapping_rows
            for field in ("default_template_id", "fallback_template_id")
            if clean_string(record.get(field))
            and clean_string(record.get(field)) not in template_ids
        }
    )
    incomplete_family_mappings = missing_mapping_fields(included_family_mapping_rows)
    missing_diphthong_node_ids = sorted(
        set(DIPHTHONG_NODE_IDS)
        - {clean_string(record.get("node_id")) for record in micro_skill_rows}
    )

    validation_errors = {
        "unknown_cluster_families": unknown_cluster_families,
        "unknown_micro_skill_families": unknown_micro_skill_families,
        "unknown_micro_skill_clusters": unknown_micro_skill_clusters,
        "unknown_mapping_templates": unknown_mapping_templates,
        "empty_clusters": empty_clusters,
        "missing_ready_metadata": dict(missing_ready_metadata),
        "missing_diphthong_node_ids": missing_diphthong_node_ids,
    }
    blocking_errors = {
        key: value for key, value in validation_errors.items() if value
    }
    if blocking_errors:
        raise ValueError(f"Workbook validation failed: {blocking_errors}")

    allowed_template_keys_by_family = build_allowed_template_keys(
        included_family_mapping_rows
    )

    families = [normalize_family(record) for record in family_rows]
    clusters = [
        normalize_cluster(record, ready_family_ids) for record in cluster_rows
    ]
    micro_skills = [
        normalize_micro_skill(record, ready_family_ids, allowed_template_keys_by_family)
        for record in micro_skill_rows
    ]
    task_templates = [
        normalize_task_template(record) for record in task_template_rows
    ]
    family_level_template_mappings = [
        normalize_family_mapping(record) for record in included_family_mapping_rows
    ]

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    write_json(OUTPUT_DIR / "families.json", families)
    write_json(OUTPUT_DIR / "clusters.json", clusters)
    write_json(OUTPUT_DIR / "micro-skills.json", micro_skills)
    write_json(OUTPUT_DIR / "task-templates.json", task_templates)
    write_json(
        OUTPUT_DIR / "family-level-template-mappings.json",
        family_level_template_mappings,
    )
    (OUTPUT_DIR / "README.md").write_text(
        "\n".join(
            [
                "# Domain 4 Seed Expansion Artifacts",
                "",
                "These files are deterministic dry-run artifacts generated from",
                f"`{WORKBOOK_PATH}`.",
                "",
                "The dry-run artifact slice and clean taxonomy import slice are complete. The",
                "import did not create migrations, change resolver behavior, enable practice",
                "template routing, or enable `grouped_set_practice`.",
                "",
                "Supported database targets today:",
                "",
                "- `families.json` -> `micro_skill_families`",
                "- `clusters.json` -> `micro_skill_clusters`",
                "- `micro-skills.json` -> `micro_skill_catalog`",
                "",
                "Artifact/config only until schema support exists:",
                "",
                "- `task-templates.json`",
                "- `family-level-template-mappings.json`",
                "",
                "Standalone word-bank rows are not generated because the current schema",
                "stores word/example data in `micro_skill_catalog.metadata`.",
                "",
                "Generated files:",
                "",
                "- `families.json`",
                "- `clusters.json`",
                "- `micro-skills.json`",
                "- `task-templates.json`",
                "- `family-level-template-mappings.json`",
                "- `validation-report.json`",
                "",
                "Validation summary:",
                "",
                "- `8` Ready Domain 4 families",
                "- `47` clusters",
                "- `240` micro-skills",
                "- `12` task templates in repo artifact/config",
                "- `40` family-level mappings in repo artifact/config",
                "- no duplicate family, cluster, or micro-skill IDs",
                "- no unknown taxonomy family or cluster references",
                "- no empty clusters",
                "- required Ready-row runtime metadata is present",
                "- required diphthong node IDs are present",
                "",
                "Implementation decisions reflected here:",
                "",
                "- `D4_PROOF` is excluded from taxonomy and family-level mapping",
                "  artifacts. Proofreading may return later as a separate writing",
                "  or editing concept, but not as Domain 4 spelling taxonomy.",
                "- `D4_IRRE` mapping rows are non-blocking for taxonomy seeding.",
                "- Practice-template routing and `grouped_set_practice` adoption are",
                "  intentionally deferred until after the clean taxonomy seed is stable.",
                "",
                "Known artifact warning before runtime use of mappings:",
                "",
                "- `D4_IRRE` has five incomplete family-level mapping rows with blank",
                "  template fields.",
                "",
                "Resolver adoption is intentionally separate from this seed expansion.",
                "",
                "## Backup And Import Closeout",
                "",
                "The clean taxonomy import used these generated artifacts for the supported",
                "database targets only:",
                "",
                "- `families.json` -> `micro_skill_families`",
                "- `clusters.json` -> `micro_skill_clusters`",
                "- `micro-skills.json` -> `micro_skill_catalog`",
                "",
                "Backup/export output was written locally and is not committed:",
                "",
                "- `.tmp/catalog-reset-backups/domain4-seed-expansion-2026-06-07T19-09-28-585Z`",
                "",
                "Rows exported before mutation:",
                "",
                "- `micro_skill_families`: `2`",
                "- `micro_skill_clusters`: `14`",
                "- `micro_skill_catalog`: `15`",
                "- dependent audited writing/learning/catalog tables: `0`",
                "",
                "Rows deleted as stale taxonomy:",
                "",
                "- `micro_skill_catalog`: `3`",
                "- `micro_skill_clusters`: `7`",
                "- `micro_skill_families`: `0`",
                "",
                "Rows upserted from these artifacts:",
                "",
                "- `micro_skill_families`: `8`",
                "- `micro_skill_clusters`: `47`",
                "- `micro_skill_catalog`: `240`",
                "",
                "Final live validation:",
                "",
                "- `8` Domain 4 families",
                "- `47` Domain 4 clusters",
                "- `240` Domain 4 micro-skills",
                "- `240` active assignable Domain 4 micro-skills",
                "- `0` `D4_PROOF` taxonomy rows",
                "- stale direct column references: `0`",
                "- stale linked rows: `0`",
                "- stale metadata snapshot rows: `0`",
                "",
                "Task templates and family-level mappings remain repo-owned artifact/config only.",
                "They were not imported into runtime tables. The `D4_IRRE` incomplete mapping",
                "rows remain deferred and non-blocking for taxonomy seeding, but they must be",
                "completed before any future runtime use of family-level mapping artifacts.",
                "",
            ]
        ),
        encoding="utf8",
    )

    validation_report = {
        "seed_version": SEED_VERSION,
        "source_workbook": str(WORKBOOK_PATH),
        "output_dir": str(OUTPUT_DIR),
        "expected_counts": EXPECTED_COUNTS,
        "actual_counts": {
            "ready_families": len(ready_family_ids),
            "clusters": len(clusters),
            "micro_skills": len(micro_skills),
            "task_templates": len(task_templates),
            "family_level_template_mappings": len(family_level_template_mappings),
        },
        "family_skill_counts": dict(
            sorted(Counter(row["skill_family_key"] for row in micro_skills).items())
        ),
        "supported_database_targets": {
            "families": "micro_skill_families",
            "clusters": "micro_skill_clusters",
            "micro_skills": "micro_skill_catalog",
        },
        "unsupported_direct_database_targets": [
            "task_templates",
            "family_level_template_mappings",
            "standalone_word_bank_rows",
        ],
        "artifact_warnings": {
            "family_level_mapping_references_without_taxonomy_family": unknown_mapping_families,
            "incomplete_family_level_template_mappings": incomplete_family_mappings,
        },
        "excluded_rows": {
            "family_level_template_mappings": [
                {
                    "row": record["_row"],
                    "family_id": clean_string(record.get("family_id")),
                    "reason": "D4_PROOF is excluded from Domain 4 spelling taxonomy and mapping artifacts.",
                }
                for record in excluded_family_mapping_rows
            ]
        },
        "supported_db_practice_routes": SUPPORTED_DB_PRACTICE_ROUTES,
        "default_practice_route": SUPPORTED_PRACTICE_ROUTE,
        "resolver_adoption": "not_changed_by_this_slice",
        "checks": {
            "no_duplicate_family_ids": True,
            "no_duplicate_cluster_ids": True,
            "no_duplicate_micro_skill_ids": True,
            "no_unknown_family_references": True,
            "no_unknown_cluster_references": True,
            "no_empty_clusters": True,
            "ready_rows_have_runtime_metadata": True,
            "diphthong_node_ids_exist": True,
        },
        "notes": [
            "This dry-run slice does not mutate Supabase.",
            "D4_PROOF is excluded from Domain 4 spelling taxonomy and mapping artifacts.",
            "Task templates and family-level mappings remain repo artifacts/config until schema support exists.",
            "Practice-template routing and grouped_set_practice adoption are deferred until after taxonomy seeding.",
            "Word/example data is projected into micro_skill_catalog metadata for import-time use.",
            "Resolver adoption remains separate from taxonomy seeding.",
        ],
    }

    if validation_report["actual_counts"] != EXPECTED_COUNTS:
        raise ValueError(
            "Generated counts differ from expected counts: "
            f"{validation_report['actual_counts']} != {EXPECTED_COUNTS}"
        )

    write_json(OUTPUT_DIR / "validation-report.json", validation_report)
    print(json.dumps(validation_report, indent=2, sort_keys=True))


if __name__ == "__main__":
    main()

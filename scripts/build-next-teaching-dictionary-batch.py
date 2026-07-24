#!/usr/bin/env python3
"""Build a deterministic, fail-closed Teaching Dictionary review batch.

The output is candidate review data only. AI-assisted mapping and dictation rows
remain ``in_review`` and cannot be promoted by the canonical CSV validator until
a named human completes the review register.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import statistics
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from wordfreq import zipf_frequency


ROOT = Path(__file__).resolve().parents[1]
INTAKE = ROOT / "docs/implementation/seed-data/teaching-dictionary/candidates/2026-06-29-phase-5-source-intake"
UK_AGES = INTAKE / "uk_spelling_words_age_estimates.csv"
BNC = INTAKE / "british_frequency_bnc_1_2_all_freq_source.txt"
AOA = INTAKE / "brysbaert_biemiller_test_based_aoa_master_source.xlsx"
D4 = ROOT / "docs/implementation/seed-data/domain4-seed-expansion/micro-skills.json"
COMMON = ROOT / "docs/implementation/seed-data/common_misspellings_seed_v1.csv"
WORD_MAP = ROOT / "docs/implementation/seed-data/canonical-spelling-word-map/canonical-spelling-word-map-v1.xlsx"

FORCED = {
    "belief": ["D4_IRRE_TRICKY_WORDS_COMPLEX_HIGH_FREQUENCY"],
    "brownie": ["D4_PG_LONG_EE_IE"],
    "diabetes": ["D4_MOR_ROOTS_COMMON_GREEK_ROOTS", "D4_MOR_ROOTS_SCIENCE_MATH_ROOTS"],
    "either": ["D4_PG_LONG_EE_EI"],
    "immigrants": ["D4_MOR_BASE_WORDS_BASE_PLUS_PREFIX"],
    "summary": ["D4_SCHWA_MEDIAL_COMMON_WEAK_VOWELS"],
}

EXISTING_REPAIRS = [
    ("business", "support_add", "D4_IRRE_TRICKY_WORDS_COMMON_HIGH_FREQUENCY"),
    ("enough", "support_add", "D4_IRRE_TRICKY_WORDS_COMMON_HIGH_FREQUENCY"),
    ("fly", "support_review", "D4_PG_LONG_IGH_Y_FINAL"),
    ("government", "support_add", "D4_MOR_BASE_WORDS_BASE_PLUS_SUFFIX"),
    ("let's", "support_add", "D4_PAT_CONTRACTIONS_PRONOUN_VERB"),
    ("govern", "metadata_add", ""),
    ("governor", "metadata_add", ""),
    ("tall", "metadata_add", ""),
    ("sign", "readiness_support", "D4_MOR_WORD_FAMILIES_RELATED_WORD_SUPPORT"),
    ("signature", "readiness_support", "D4_MOR_WORD_FAMILIES_RELATED_WORD_SUPPORT"),
    ("bomb", "readiness_support", "D4_MOR_WORD_FAMILIES_RELATED_WORD_SUPPORT"),
    ("bombard", "readiness_support", "D4_MOR_WORD_FAMILIES_RELATED_WORD_SUPPORT"),
    ("nature", "readiness_support", "D4_MOR_WORD_FAMILIES_RELATED_WORD_SUPPORT"),
    ("natural", "readiness_support", "D4_MOR_WORD_FAMILIES_RELATED_WORD_SUPPORT"),
    ("heal", "readiness_support", "D4_MOR_WORD_FAMILIES_RELATED_WORD_SUPPORT"),
    ("health", "readiness_support", "D4_MOR_WORD_FAMILIES_RELATED_WORD_SUPPORT"),
]

CSV_HEADERS = {
    "canonical_words.csv": ["word_key", "normalised_word", "display_word", "dialect_code", "frequency_band", "age_band", "complexity_band", "source_category", "source_name", "source_url", "source_licence", "source_use_note", "confidence", "review_status", "row_status"],
    "canonical_word_metadata.csv": ["word_key", "syllables", "phoneme_hint", "grapheme_notes", "stress_pattern", "has_schwa", "morphemes", "morphology_notes", "irregularity_notes", "source_category", "source_name", "source_url", "source_licence", "source_use_note", "confidence", "review_status"],
    "micro_skill_word_support.csv": ["word_key", "micro_skill_key", "support_role", "source_category", "source_name", "source_url", "source_licence", "source_use_note", "confidence", "review_status", "review_notes"],
    "dictation_sentences.csv": ["word_key", "display_word", "age_band", "complexity_band", "dictation_sentence", "dictation_target_token_index", "audio_text", "source_category", "source_name", "source_url", "source_licence", "source_use_note", "confidence", "review_status", "reviewed_by", "reviewed_at", "review_notes"],
    "teaching_content_versions.csv": ["micro_skill_key", "content_version", "version_status", "is_active", "teaching_objective", "child_friendly_explanation", "rule_explanation", "memory_tip", "common_misconceptions", "first_exposure_progression", "guided_practice_progression", "review_proofreading_progression", "example_selection_guidance", "contrast_policy_guidance", "sample_preview_word_key", "source_category", "source_name", "source_url", "source_licence", "source_use_note", "confidence", "supersedes_content_version", "final_readiness_review_status", "final_readiness_reviewed_by", "final_readiness_reviewed_at"],
    "teaching_content_field_reviews.csv": ["micro_skill_key", "content_version", "field_key", "review_gate", "review_status", "reviewed_by", "reviewed_at", "review_notes"],
    "teaching_content_sources.csv": ["source_key", "source_category", "source_name", "source_url", "source_licence", "source_use_note", "importability_status", "legal_review_status"],
}


def norm(value: Any) -> str:
    return str(value or "").strip().lower().replace("’", "'")


def key_for(word: str) -> str:
    return f"{re.sub(r'[^a-z0-9]+', '_', word).strip('_')}_en_gb"


def read_csv(path: Path, delimiter: str = ",") -> list[dict[str, str]]:
    with path.open(encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle, delimiter=delimiter))


def write_csv(path: Path, headers: list[str], rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def age_band(age: int) -> str:
    if age <= 7: return "early_primary"
    if age <= 9: return "middle_primary"
    if age <= 11: return "upper_primary"
    if age <= 13: return "lower_secondary"
    if age <= 15: return "mid_secondary"
    return "later_review"


def frequency_band(zipf: float) -> str:
    return "high" if zipf >= 4 else "medium" if zipf >= 3 else "low"


def bnc_index() -> dict[str, float]:
    totals: dict[str, float] = defaultdict(float)
    for row in read_csv(BNC, "\t"):
        word = norm(row.get("Word", "")).removesuffix("*")
        if word:
            try: totals[word] += float(row.get("Freq", "") or 0)
            except ValueError: pass
    return totals


def aoa_index() -> dict[str, float]:
    workbook = load_workbook(AOA, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = sheet.iter_rows(values_only=True)
    headers = [str(v or "") for v in next(rows)]
    word_i, age_i = headers.index("WORD"), headers.index("AoAtestbased")
    values: dict[str, list[float]] = defaultdict(list)
    for row in rows:
        word = norm(row[word_i])
        try: age = float(row[age_i])
        except (TypeError, ValueError): continue
        if word and " " not in word and "(" not in word:
            values[word].append(age)
    return {word: statistics.median(ages) for word, ages in values.items()}


def ipa_index(path: Path) -> dict[str, str]:
    result = {}
    for line in path.read_text(encoding="utf-8").splitlines():
        if "\t" not in line: continue
        word, ipa = line.split("\t", 1)
        result.setdefault(norm(word), ipa.strip())
    return result


def cmu_index(path: Path) -> dict[str, str]:
    result = {}
    for line in path.read_text(encoding="latin-1").splitlines():
        if not line or line.startswith(";;; ") or " " not in line: continue
        word, phones = line.split(" ", 1)
        result.setdefault(norm(re.sub(r"\(\d+\)$", "", word)), phones.strip())
    return result


def morpholex_index(path: Path) -> dict[str, dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    result: dict[str, dict[str, Any]] = {}
    for name in workbook.sheetnames:
        if not re.fullmatch(r"\d+-\d+-\d+", name): continue
        rows = workbook[name].iter_rows(values_only=True)
        headers = [str(v or "") for v in next(rows)]
        if "Word" not in headers or "MorphoLexSegm" not in headers: continue
        indexes = {header: i for i, header in enumerate(headers)}
        for row in rows:
            word = norm(row[indexes["Word"]])
            if not word or word in result: continue
            prs = str(row[indexes.get("PRS_signature", 0)] or name).replace(",", "-")
            nums = [int(v) for v in re.findall(r"\d+", prs)[:3]]
            result[word] = {
                "segmentation": str(row[indexes["MorphoLexSegm"]] or ""),
                "pos": str(row[indexes.get("POS", 0)] or ""),
                "nmorph": int(row[indexes.get("Nmorph", 0)] or 1),
                "prefixes": nums[0] if len(nums) == 3 else 0,
                "suffixes": nums[2] if len(nums) == 3 else 0,
            }
    return result


def prior_source_words() -> set[str]:
    words = {norm(row.get("correction")) for row in read_csv(COMMON)}
    workbook = load_workbook(WORD_MAP, read_only=True, data_only=True)
    for sheet_name in ("micro_skill_word_bank", "word_metadata", "contrast_pairs", "diagnostic_misspelling_mappings"):
        sheet = workbook[sheet_name]
        rows = sheet.iter_rows(values_only=True)
        headers = [str(v or "") for v in next(rows)]
        for row in rows:
            record = dict(zip(headers, row))
            for field in ("normalised_word", "target_word", "contrast_word", "correction_normalised"):
                if record.get(field): words.add(norm(record[field]))
    for skill in json.loads(D4.read_text(encoding="utf-8")):
        meta = skill.get("metadata", {})
        for value in meta.get("example_words", []):
            words.update(norm(part) for part in re.split(r"[/\s]+", str(value)) if part)
        for row in meta.get("starter_word_bank", []):
            words.update(norm(part) for part in re.split(r"[/\s]+", str(row.get("word", ""))) if part)
    return {word for word in words if word}


def syllable_facts(ipa: str, cmu: str) -> tuple[int, str, bool]:
    stresses = re.findall(r"[012]", cmu)
    if stresses:
        labels = {"0": "unstressed", "1": "primary", "2": "secondary"}
        return len(stresses), "-".join(labels[s] for s in stresses), "AH0" in cmu or "ER0" in cmu
    nuclei = re.findall(r"[aeiouæɑɒɔəɜɛɪʊʌɐ]+", ipa.replace("ː", ""))
    return max(1, len(nuclei)), "in_review", any(symbol in ipa for symbol in ("ə", "ɐ"))


def proposed_skill(word: str, zipf: float, morph: dict[str, Any], catalog: set[str]) -> str:
    if word in FORCED: return FORCED[word][0]
    non_morphology_rules = [
        (r"'", "D4_PAT_CONTRACTIONS_PRONOUN_VERB"),
        (r"^(kn|wr)", "D4_PAT_SILENT_LETTERS_KN" if word.startswith("kn") else "D4_PAT_SILENT_LETTERS_WR"),
        (r"ould$", "D4_IRRE_TRICKY_WORDS_OULD_WORDS"),
        (r"ough", "D4_IRRE_TRICKY_WORDS_COMPLEX_HIGH_FREQUENCY"),
        (r"ie", "D4_PG_LONG_EE_IE"),
        (r"ei", "D4_PG_LONG_EE_EI"),
    ]
    prefix_rules = [
        (r"^un", "D4_MOR_PREFIXES_UN"),
        (r"^(re|pre)", "D4_MOR_PREFIXES_RE_PRE"),
        (r"^(dis|mis)", "D4_MOR_PREFIXES_DIS_MIS"),
        (r"^(in|im|il|ir)", "D4_MOR_PREFIXES_IN_IM_IL_IR"),
        (r"^(sub|inter|super)", "D4_MOR_PREFIXES_SUB_INTER_SUPER"),
    ]
    suffix_rules = [
        (r"ness$", "D4_MOR_SUFFIXES_NESS"),
        (r"(ful|less)$", "D4_MOR_SUFFIXES_FUL_LESS"),
        (r"ment$", "D4_MOR_SUFFIXES_MENT"),
        (r"ous$", "D4_MOR_SUFFIXES_OUS"),
        (r"ity$", "D4_MOR_SUFFIXES_ITY"),
        (r"al$", "D4_MOR_SUFFIXES_AL"),
        (r"(able|ible)$", "D4_MOR_SUFFIXES_ABLE_IBLE"),
        (r"ing$", "D4_INF_ING_ENDINGS_REGULAR"),
        (r"ied$", "D4_INF_PAST_TENSE_ED_Y_TO_I"),
        (r"ed$", "D4_INF_PAST_TENSE_ED_REGULAR"),
        (r"ies$", "D4_INF_PLURALS_IES"),
        (r"ves$", "D4_INF_PLURALS_VES"),
        (r"(ches|shes|sses|xes|zes)$", "D4_INF_PLURALS_ES"),
    ]
    rules = list(non_morphology_rules)
    if morph.get("prefixes", 0): rules.extend(prefix_rules)
    if morph.get("suffixes", 0): rules.extend(suffix_rules)
    for pattern, skill in rules:
        if re.search(pattern, word) and skill in catalog:
            return skill
    if morph.get("prefixes", 0): return "D4_MOR_BASE_WORDS_BASE_PLUS_PREFIX"
    if morph.get("suffixes", 0): return "D4_MOR_BASE_WORDS_BASE_PLUS_SUFFIX"
    if zipf >= 4.5 and len(word) <= 5: return "D4_IRRE_TRICKY_WORDS_SIMPLE_HIGH_FREQUENCY"
    if zipf >= 4: return "D4_IRRE_TRICKY_WORDS_COMMON_HIGH_FREQUENCY"
    return "D4_IRRE_TRICKY_WORDS_COMPLEX_HIGH_FREQUENCY"


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--snapshot", type=Path, required=True)
    parser.add_argument("--ipa", type=Path, required=True)
    parser.add_argument("--cmudict", type=Path, required=True)
    parser.add_argument("--morpholex", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--limit", type=int, default=1000)
    args = parser.parse_args()
    if args.limit != 1000: raise SystemExit("The accepted batch contract requires --limit 1000.")

    snapshot = json.loads(args.snapshot.read_text(encoding="utf-8"))
    active = {row["normalised_word"] for row in snapshot["activeWords"]}
    curriculum = read_csv(UK_AGES)
    bnc = bnc_index()
    aoa = aoa_index()
    ipa = ipa_index(args.ipa)
    cmu = cmu_index(args.cmudict)
    morph = morpholex_index(args.morpholex)
    prior = prior_source_words()
    catalog_rows = json.loads(D4.read_text(encoding="utf-8"))
    catalog = {row["micro_skill_key"] for row in catalog_rows if row.get("is_active")}

    candidates = []
    for row in curriculum:
        display = str(row["Word"]).strip()
        word = norm(display)
        if not word or word in active: continue
        zipf = round(float(zipf_frequency(word, "en")), 3)
        evidence_age = int(row["Age"])
        m = morph.get(word, {})
        candidates.append({
            "word": word, "display": display, "curriculum_age": evidence_age,
            "zipf": zipf, "bnc_frequency": bnc.get(word, 0), "aoa": aoa.get(word),
            "prior_source": word in prior, "morph": m,
            "forced": word in FORCED,
        })

    missing_forced = sorted(set(FORCED) - {row["word"] for row in candidates} - active)
    for word in missing_forced:
        zipf = round(float(zipf_frequency(word, "en")), 3)
        candidates.append({"word": word, "display": word, "curriculum_age": int(round(aoa.get(word, 10))), "zipf": zipf, "bnc_frequency": bnc.get(word, 0), "aoa": aoa.get(word), "prior_source": word in prior, "morph": morph.get(word, {}), "forced": True})

    candidates.sort(key=lambda row: (
        0 if row["forced"] else 1,
        -row["zipf"],
        0 if row["prior_source"] else 1,
        0 if row["morph"].get("prefixes", 0) or row["morph"].get("suffixes", 0) else 1,
        row["curriculum_age"], row["word"],
    ))
    selected = candidates[: args.limit]
    if len(selected) != args.limit: raise SystemExit(f"Only {len(selected)} eligible candidates were found.")
    if not set(FORCED).issubset({row["word"] for row in selected}): raise SystemExit("Forced learner-demand words were not all selected.")

    source_note = "Candidate selected from the project-authored UK spelling-age evidence list; human review required before promotion."
    words_rows, metadata_rows, support_rows, dictation_rows, register = [], [], [], [], []
    for rank, row in enumerate(selected, 1):
        word, display, m = row["word"], row["display"], row["morph"]
        word_key = key_for(word)
        ipa_value, cmu_value = ipa.get(word, ""), cmu.get(word, "")
        syllables, stress, has_schwa = syllable_facts(ipa_value, cmu_value)
        age_value = int(round(row["aoa"])) if row["aoa"] is not None else row["curriculum_age"]
        complexity_score = len(word) + syllables * 2 + int(m.get("nmorph", 1))
        complexity = "low" if complexity_score <= 9 else "medium" if complexity_score <= 15 else "high"
        skill = proposed_skill(word, row["zipf"], m, catalog)
        skills = FORCED.get(word, [skill])
        morphology_verified = bool(m)
        source_category = "internal_authored"
        words_rows.append({"word_key": word_key, "normalised_word": word, "display_word": display, "dialect_code": "en-GB", "frequency_band": frequency_band(row["zipf"]), "age_band": age_band(age_value), "complexity_band": complexity, "source_category": source_category, "source_name": "Next Teaching Dictionary curriculum-evidence intake", "source_url": str(UK_AGES.relative_to(ROOT)), "source_licence": "internal/project-authored", "source_use_note": source_note, "confidence": "medium", "review_status": "in_review", "row_status": "draft"})
        metadata_rows.append({"word_key": word_key, "syllables": syllables, "phoneme_hint": ipa_value or cmu_value, "grapheme_notes": "", "stress_pattern": stress, "has_schwa": "TRUE" if has_schwa else "FALSE", "morphemes": m.get("segmentation", ""), "morphology_notes": f"MorphoLex-en exact candidate: {m.get('segmentation', '')}; POS={m.get('pos', '')}." if m else "No exact MorphoLex-en row; human morphology review required.", "irregularity_notes": "AI-assisted candidate metadata; verify British pronunciation and teaching relevance.", "source_category": "ai_assisted_draft", "source_name": "British IPA, CMUdict and MorphoLex candidate enrichment", "source_url": "https://github.com/open-dict-data/ipa-dict; https://github.com/cmusphinx/cmudict; https://github.com/hugomailhot/MorphoLex-en", "source_licence": "mixed open licences; see teaching_content_sources.csv", "source_use_note": "Automated candidate enrichment only; not approved teaching truth.", "confidence": "medium" if ipa_value and cmu_value and morphology_verified else "low", "review_status": "in_review"})
        for proposed in skills:
            support_rows.append({"word_key": word_key, "micro_skill_key": proposed, "support_role": "support_example", "source_category": "ai_assisted_draft", "source_name": "Next Teaching Dictionary proposed mapping", "source_url": str(D4.relative_to(ROOT)), "source_licence": "internal", "source_use_note": "Deterministic proposed mapping; requires human pedagogy review.", "confidence": "medium" if word in FORCED or morphology_verified else "low", "review_status": "in_review", "review_notes": "Confirm that this word genuinely teaches the proposed micro-skill; do not approve merely to satisfy a learning-item gap."})
        sentence = f"The class practised the word {display}."
        dictation_rows.append({"word_key": word_key, "display_word": display, "age_band": age_band(age_value), "complexity_band": complexity, "dictation_sentence": sentence, "dictation_target_token_index": 5, "audio_text": sentence, "source_category": "ai_assisted_draft", "source_name": "Next Teaching Dictionary draft dictation", "source_url": "", "source_licence": "internal", "source_use_note": "Placeholder draft requiring semantic, child-language, accessibility and British-English review.", "confidence": "low", "review_status": "in_review", "reviewed_by": "", "reviewed_at": "", "review_notes": "Replace generic placeholder with a contextual sentence before approval."})
        register.append({"rank": rank, "word_key": word_key, "word": word, "display_word": display, "forced_learner_demand": row["forced"], "curriculum_evidence": True, "curriculum_age": row["curriculum_age"], "aoa_median": row["aoa"] or "", "wordfreq_zipf": row["zipf"], "frequency_band": frequency_band(row["zipf"]), "bnc_frequency": row["bnc_frequency"], "prior_source_continuity": row["prior_source"], "morpholex_exact": morphology_verified, "verified_prefix_count": m.get("prefixes", 0), "verified_suffix_count": m.get("suffixes", 0), "morpholex_segmentation": m.get("segmentation", ""), "british_ipa": ipa_value, "cmudict": cmu_value, "syllables": syllables, "age_band": age_band(age_value), "complexity_band": complexity, "proposed_micro_skill_keys": "|".join(skills), "mapping_review": "in_review", "morphology_review": "in_review", "pronunciation_review": "in_review", "dictation_review": "in_review", "british_english_review": "in_review", "accessibility_review": "in_review", "source_licence_review": "in_review", "final_decision": "in_review", "reviewed_by": "", "reviewed_at": "", "review_notes": ""})

    csv_dir = args.output / "csv"
    payloads = {"canonical_words.csv": words_rows, "canonical_word_metadata.csv": metadata_rows, "micro_skill_word_support.csv": support_rows, "dictation_sentences.csv": dictation_rows, "teaching_content_versions.csv": [], "teaching_content_field_reviews.csv": [], "teaching_content_sources.csv": [
        {"source_key": "next_batch_uk_curriculum_evidence", "source_category": "internal_authored", "source_name": "UK spelling words age estimates", "source_url": str(UK_AGES.relative_to(ROOT)), "source_licence": "internal/project-authored", "source_use_note": source_note, "importability_status": "importable", "legal_review_status": "passed"},
        {"source_key": "next_batch_wordfreq", "source_category": "open_licensed", "source_name": "wordfreq", "source_url": "https://github.com/rspeer/wordfreq", "source_licence": "Apache-2.0 package; underlying data use remains review-gated", "source_use_note": "Selection ranking only until source review passes.", "importability_status": "requires_legal_review", "legal_review_status": "required"},
        {"source_key": "next_batch_british_ipa", "source_category": "open_licensed", "source_name": "open-dict-data ipa-dict en_UK", "source_url": "https://github.com/open-dict-data/ipa-dict", "source_licence": "MIT repository; English UK data credited under GPL-3.0", "source_use_note": "Candidate British pronunciation metadata.", "importability_status": "importable", "legal_review_status": "passed"},
        {"source_key": "next_batch_cmudict", "source_category": "open_licensed", "source_name": "CMUdict", "source_url": "https://github.com/cmusphinx/cmudict", "source_licence": "BSD-style", "source_use_note": "Pronunciation comparison/fallback evidence.", "importability_status": "importable", "legal_review_status": "passed"},
        {"source_key": "next_batch_morpholex", "source_category": "open_licensed", "source_name": "MorphoLex-en", "source_url": "https://github.com/hugomailhot/MorphoLex-en", "source_licence": "CC BY-NC-SA 4.0", "source_use_note": "Candidate morphology segmentation; human review required.", "importability_status": "importable", "legal_review_status": "passed"},
    ]}
    for filename, rows in payloads.items(): write_csv(csv_dir / filename, CSV_HEADERS[filename], rows)
    review_headers = list(register[0])
    write_csv(args.output / "review" / "selection_register.csv", review_headers, register)
    repair_rows = [{"word": word, "repair_type": repair, "micro_skill_key": skill, "review_status": "in_review", "reviewed_by": "", "reviewed_at": "", "review_notes": ""} for word, repair, skill in EXISTING_REPAIRS]
    write_csv(args.output / "review" / "existing_row_repairs.csv", list(repair_rows[0]), repair_rows)

    metadata_review = [{**row, "pronunciation_review": "in_review", "morphology_review": "in_review", "british_english_review": "in_review", "final_decision": "in_review", "reviewed_by": "", "reviewed_at": "", "review_notes": ""} for row in metadata_rows]
    mapping_review = [{**row, "mapping_review": "in_review", "final_decision": "in_review", "reviewed_by": "", "reviewed_at": "", "review_notes_final": ""} for row in support_rows]
    dictation_review = [{**row, "child_language_review": "in_review", "british_english_review": "in_review", "accessibility_review": "in_review", "final_decision": "in_review"} for row in dictation_rows]
    source_review = [{**row, "source_review": "in_review", "reviewed_by": "", "reviewed_at": "", "review_notes": ""} for row in payloads["teaching_content_sources.csv"]]
    write_csv(args.output / "review" / "metadata_review.csv", list(metadata_review[0]), metadata_review)
    write_csv(args.output / "review" / "mapping_review.csv", list(mapping_review[0]), mapping_review)
    write_csv(args.output / "review" / "dictation_review.csv", list(dictation_review[0]), dictation_review)
    write_csv(args.output / "review" / "source_review.csv", list(source_review[0]), source_review)

    workbook_data = {
        "Summary": [["Metric", "Value"], ["Batch size", len(selected)], ["Production active words", snapshot["counts"]["active_words"]], ["Expected active words after approved import", snapshot["counts"]["active_words"] + len(selected)], ["Forced learner-demand words", sum(1 for row in selected if row["forced"])], ["MorphoLex exact candidates", sum(1 for row in selected if row["morph"])], ["Status", "HUMAN REVIEW REQUIRED"]],
        "Selection register": [review_headers] + [[row[h] for h in review_headers] for row in register],
        "Metadata review": [list(metadata_review[0])] + [[row[h] for h in metadata_review[0]] for row in metadata_review],
        "Mapping review": [list(mapping_review[0])] + [[row[h] for h in mapping_review[0]] for row in mapping_review],
        "Dictation review": [list(dictation_review[0])] + [[row[h] for h in dictation_review[0]] for row in dictation_review],
        "Existing repairs": [list(repair_rows[0])] + [[row[h] for h in repair_rows[0]] for row in repair_rows],
        "Sources": [list(source_review[0])] + [[row.get(h, "") for h in source_review[0]] for row in source_review],
    }
    (args.output / "review-workbook-data.json").write_text(json.dumps(workbook_data, ensure_ascii=False), encoding="utf-8")
    manifest = {"schemaVersion": "next_teaching_dictionary_batch_v1", "batchSize": len(selected), "newWordRows": len(words_rows), "metadataRows": len(metadata_rows), "supportRows": len(support_rows), "dictationRows": len(dictation_rows), "existingRepairRows": len(repair_rows), "status": "human_review_required", "productionMutationAuthorised": False, "sourceSnapshot": str(args.snapshot), "selectionSha256": hashlib.sha256("\n".join(row["word"] for row in selected).encode()).hexdigest()}
    (args.output / "manifest.json").write_text(json.dumps(manifest, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(manifest))


if __name__ == "__main__":
    main()

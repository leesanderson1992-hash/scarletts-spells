#!/usr/bin/env python3
"""Emit approved CSVs only after complete named workbook review."""

from __future__ import annotations

import argparse
import csv
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


APPROVED = "approved"
APPROVED_GATE_VALUES = {"approved", "passed"}


def clean(value: Any) -> str:
    return "" if value is None else str(value).strip()


def rows_from_sheet(workbook: Any, name: str) -> list[dict[str, str]]:
    # Excel limits worksheet names to 31 characters; accept the deterministic
    # truncation produced by the review-workbook builder.
    sheet = workbook[name] if name in workbook.sheetnames else workbook[name[:31]]
    values = sheet.iter_rows(values_only=True)
    headers = [clean(value) for value in next(values)]
    return [dict(zip(headers, (clean(value) for value in row))) for row in values if any(clean(value) for value in row)]


def require_approved(rows: list[dict[str, str]], gates: list[str], label: str, expected: int | None = None) -> None:
    if expected is not None and len(rows) != expected:
        raise ValueError(f"{label}: expected {expected} rows, found {len(rows)}")
    failures = []
    for index, row in enumerate(rows, 2):
        missing = [gate for gate in gates if row.get(gate) not in APPROVED_GATE_VALUES]
        if not row.get("reviewed_by") or not row.get("reviewed_at"):
            missing.append("named_review")
        if missing:
            failures.append({"row": index, "word": row.get("word") or row.get("word_key") or row.get("source_key"), "missing": missing})
    if failures:
        raise ValueError(f"{label}: {len(failures)} rows are not fully approved; first failures: {json.dumps(failures[:10])}")


def write_csv(path: Path, headers: list[str], rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def headers(path: Path) -> list[str]:
    with path.open(encoding="utf-8-sig", newline="") as handle:
        return list(next(csv.reader(handle)))


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", type=Path, required=True)
    parser.add_argument("--candidate-csv", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--pronunciation-supplements", type=Path)
    args = parser.parse_args()

    workbook = load_workbook(args.workbook, read_only=True, data_only=True)
    selection = rows_from_sheet(workbook, "Canonical word review")
    morphology = rows_from_sheet(workbook, "Linguistic morphology & word sums")
    dictations = rows_from_sheet(workbook, "Dictation review")
    sources = rows_from_sheet(workbook, "Sources & licence")
    repairs = rows_from_sheet(workbook, "Existing word-fact repairs")

    require_approved(selection, ["identity_review", "selection_evidence_review", "pronunciation_review", "british_english_review", "accessibility_review", "source_licence_review", "final_decision"], "canonical words", 1000)
    require_approved(morphology, ["linguistic_analysis_review", "word_sum_review", "final_decision"], "morphology", 1000)
    require_approved(dictations, ["child_language_review", "british_english_review", "accessibility_review", "final_decision"], "dictations", 1000)
    require_approved(sources, ["source_review"], "sources")
    require_approved(repairs, ["review_status"], "existing repairs")

    generic = [row["word_key"] for row in dictations if row.get("dictation_sentence", "").startswith("The class practised the word ")]
    if generic:
        raise ValueError(f"Dictation review: {len(generic)} generic placeholders remain; first: {generic[:10]}")
    unresolved_sources = [row["source_key"] for row in sources if row.get("importability_status") != "importable" or row.get("legal_review_status") not in {"passed", "not_required"}]
    if unresolved_sources:
        raise ValueError(f"Source review remains blocked: {unresolved_sources}")

    words_path = args.candidate_csv / "canonical_words.csv"
    with words_path.open(encoding="utf-8-sig", newline="") as handle:
        words = list(csv.DictReader(handle))
    for row in words:
        row["row_status"] = "active"
        row["review_status"] = "approved_for_first_exposure"
        row["confidence"] = "high"
    valid_decisions = {"approved", "not_applicable", "rejected"}
    unresolved = [row["word_key"] for row in morphology if row.get("analysis_status") not in valid_decisions]
    approved_without_sum = [row["word_key"] for row in morphology if row.get("analysis_status") == "approved" and (not row.get("word_sum") or row.get("morphology_parts") in {"", "[]"})]
    if unresolved or approved_without_sum:
        raise ValueError(f"Morphology decisions are not import-safe: unresolved={unresolved[:10]}, approved_without_word_sum={approved_without_sum[:10]}")
    for row in morphology:
        row["review_status"] = "approved_for_first_exposure"
        row["confidence"] = "high"
    for row in dictations:
        row["review_status"] = "approved_for_first_exposure"
        row["confidence"] = "high"
        # Workbook-facing labels may be more specific than the persisted
        # provenance enum; retain the authored provenance in source_name/note.
        if row.get("source_category") == "project_authored_short_context":
            row["source_category"] = "internal_authored"

    write_csv(args.output / "canonical_words.csv", headers(words_path), words)
    # Metadata are factual candidate enrichment; promotion retains the reviewed source rows.
    with (args.candidate_csv / "canonical_word_metadata.csv").open(encoding="utf-8-sig", newline="") as handle:
        metadata = list(csv.DictReader(handle))
    selection_by_key = {row["word_key"]: row for row in selection}
    supplement_path = args.pronunciation_supplements or args.candidate_csv.parent / "review" / "reviewed-pronunciation-supplements.json"
    supplements: dict[str, dict[str, str]] = {}
    supplement_document: dict[str, Any] = {}
    if supplement_path.exists():
        supplement_document = json.loads(supplement_path.read_text(encoding="utf-8"))
        if supplement_document.get("schemaVersion") != "reviewed_pronunciation_supplements_v1":
            raise ValueError("Pronunciation supplements use an unsupported schema.")
        supplements = {row["word_key"]: row for row in supplement_document.get("rows", [])}
    for row in metadata:
        word_key = row["word_key"]
        selection_row = selection_by_key[word_key]
        reviewed_ipa = selection_row.get("british_ipa", "")
        if reviewed_ipa:
            row["phoneme_hint"] = reviewed_ipa
        supplement = supplements.get(word_key)
        if supplement:
            if supplement.get("british_ipa") != reviewed_ipa:
                raise ValueError(f"{word_key}: pronunciation supplement does not match the approved workbook IPA.")
            row.update(
                {
                    "syllables": supplement["syllables"],
                    "phoneme_hint": supplement["british_ipa"],
                    "stress_pattern": supplement["stress_pattern"],
                    "has_schwa": supplement["has_schwa"],
                    "source_category": "internal_authored",
                    "source_name": "Katie Sanderson reviewed pronunciation supplement",
                    "source_url": str(supplement_path),
                    "source_licence": "internal",
                    "source_use_note": supplement_document["sourceNote"],
                }
            )
        row.update({"review_status": "approved_for_first_exposure", "confidence": "high"})
        if (
            not row.get("phoneme_hint")
            or not row.get("syllables")
            or row.get("stress_pattern") in {"", "in_review", "unknown"}
            or row.get("has_schwa") not in {"TRUE", "FALSE"}
        ):
            raise ValueError(f"{word_key}: approved pronunciation metadata remains incomplete.")
    write_csv(args.output / "canonical_word_metadata.csv", headers(args.candidate_csv / "canonical_word_metadata.csv"), metadata)
    morphology_headers = headers(args.candidate_csv / "canonical_word_morphology.csv")
    write_csv(args.output / "canonical_word_morphology.csv", morphology_headers, morphology)
    write_csv(args.output / "dictation_sentences.csv", headers(args.candidate_csv / "dictation_sentences.csv"), dictations)
    write_csv(args.output / "teaching_content_sources.csv", headers(args.candidate_csv / "teaching_content_sources.csv"), sources)
    (args.output.parent / "approved-existing-row-repairs.json").write_text(json.dumps(repairs, indent=2) + "\n", encoding="utf-8")
    print(json.dumps({"status": "approved_csv_emitted", "words": len(words), "metadata": len(metadata), "morphology": len(morphology), "support_links": 0, "dictations": len(dictations), "repairs": len(repairs)}))


if __name__ == "__main__":
    main()

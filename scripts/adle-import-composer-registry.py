#!/usr/bin/env python3
"""ADLE Slice 3 (3B): guarded importer for the composer registry sheets.

Reads the content workbook's Family Methods (8 rows) and Activity Templates
(32 rows) sheets and lands them as reviewable content data in
adle_family_methods / adle_activity_templates through the established
import-batch pattern. Dry-run is the default; --apply is guarded to local/dev
exactly like scripts/adle-band-teaching-dictionary.py.

Policy boundaries (adle-slice-3-daily-assignment-composer-plan.md, 3B):
  - these two sheets are content/registry data; the workbook's policy columns
    are never read (the Micro Skill Content sheet is not touched — its content
    is already imported as canonical_teaching_dictionary_content_versions)
  - guided sequences may reference the two documented composition-time
    meta-keys (DICTATION_OR_WRITING, SENTENCE_APPLICATION); every other
    sequence key must resolve to an Activity Templates row
  - the registry contract's runtime metadata (min_words_required,
    requires_sentence_context, requires_contrast_words, evidence_kind) is
    declared here per template key and validated to cover the sheet exactly;
    evidence_kind is a label only — weights are Slice 4's
  - re-import supersedes prior active rows and inserts under a new batch
"""

from __future__ import annotations

import argparse
import json
import subprocess
import sys
from pathlib import Path
from typing import Any
from urllib.parse import urlparse

ROOT = Path(__file__).resolve().parents[1]

CONTENT_VERSION = "2026-07-04.v1"
IMPORTER_VERSION = "adle_composer_registry_importer_v1"
DEFAULT_WORKBOOK = ROOT / "docs/implementation/seed-data/ADLE_content_workbook_v1.xlsx"

FAMILY_SHEET = "Family Methods"
TEMPLATE_SHEET = "Activity Templates"
FAMILY_HEADERS = (
    "family_key", "family_name", "core_pedagogy", "first_exposure_sequence",
    "guided_question_sequence", "review_sort_dimension", "production_task", "notes",
)
TEMPLATE_HEADERS = (
    "template_key", "phase", "purpose", "child_response", "required_inputs",
    "child_facing_copy",
)

EXPECTED_FAMILY_COUNT = 8
EXPECTED_TEMPLATE_COUNT = 32

# The taxonomy's skill_family_key set (micro_skill_catalog); apply mode
# re-checks this against the live catalog so drift fails the batch.
EXPECTED_FAMILY_KEYS = frozenset({
    "D4_PG", "D4_PAT", "D4_SYL", "D4_HOM", "D4_IRRE", "D4_MOR", "D4_INF", "D4_SCHWA",
})

# Composition-time meta-keys allowed inside guided sequences: the composer
# resolves them to production templates (dictation / must-use writing /
# sentence-context dictation) per the family's production task.
SEQUENCE_META_KEYS = frozenset({"DICTATION_OR_WRITING", "SENTENCE_APPLICATION"})

# Registry-contract runtime metadata per template key
# (min_words_required, requires_sentence_context, requires_contrast_words,
# evidence_kind). Labels follow the blueprint's activity-set amendment;
# contrast-word requirements apply to homophone-family templates only.
TEMPLATE_RUNTIME_METADATA: dict[str, tuple[int, bool, bool, str]] = {
    "MICRO_READ_ONLY_INTRO": (1, False, False, "read_only"),
    "LESSON_WORDS_INTRO": (1, False, False, "read_only"),
    "PG_SOUND_NOTICE": (1, False, False, "guided_task"),
    "PG_GRAPHEME_MAP": (1, False, False, "guided_task"),
    "PAT_PATTERN_SPOT": (2, False, False, "guided_task"),
    "PAT_RULE_APPLY": (1, False, False, "guided_task"),
    "SYL_SPLIT": (1, False, False, "guided_task"),
    "SYL_REBUILD": (1, False, False, "guided_task"),
    "HOM_MEANING_MATCH": (1, False, True, "guided_task"),
    "HOM_SENTENCE_CHOICE": (1, True, True, "guided_task"),
    "HOM_CORRECTION": (1, True, True, "guided_task"),
    "IRRE_TRICKY_PART": (1, False, False, "guided_task"),
    "MEMORY_CUE": (1, False, False, "reflection"),
    "HIDE_WRITE": (1, False, False, "controlled_spelling"),
    "MOR_STRIP_BUILD": (1, False, False, "guided_task"),
    "MOR_MEANING_MATCH": (1, False, False, "guided_task"),
    "MOR_BUILD_WORD": (1, False, False, "guided_task"),
    "INF_CONTEXT_CHOICE": (1, True, False, "guided_task"),
    "INF_RULE_CHOICE": (1, False, False, "guided_task"),
    "INF_TRANSFORM": (1, False, False, "controlled_spelling"),
    "SCHWA_STRESS_MARK": (1, False, False, "guided_task"),
    "SCHWA_VOWEL_REVEAL": (1, False, False, "guided_task"),
    "SCHWA_ANCHOR": (1, False, False, "guided_task"),
    "CONTROLLED_SPELLING": (1, False, False, "controlled_spelling"),
    "DICTATION_NO_IMAGE": (1, False, False, "dictation"),
    "MUST_USE_FREEWRITING": (3, False, False, "free_writing"),
    "REVIEW_QUICK_SORT": (2, False, False, "categorisation"),
    "REVIEW_DICTATION": (1, False, False, "dictation"),
    "REVIEW_MUST_USE_WRITING": (3, False, False, "free_writing"),
    "ERROR_REFLECTION_CUE": (1, False, False, "reflection"),
    "DICTATION_SENTENCE_CONTEXT": (1, True, False, "dictation_sentence_context"),
    "DIAGNOSTIC_DICTATION_PROBE": (1, False, False, "diagnostic_probe"),
}

# Phases as they appear on the sheet. Unknown phases fail the batch report.
KNOWN_PHASES = frozenset({
    "Lesson intro",
    "Guided practice",
    "Transfer/proofreading",
    "Reflection/encoding",
    "Independent retrieval",
    "Encoding",
    "Independent retrieval/review",
    "Transfer/review",
    "Review interleaving",
    "Review production",
    "Review transfer",
    "Reflection",
    "Diagnostic",
})

LOCAL_CONFIRMATION_TOKEN = "adle-composer-registry-local-dev"
ADVISORY_LOCK_NAME = "adle_composer_registry_import"
EXPECTED_MIGRATION_VERSIONS = ("20260629120000", "20260705180000")

FAMILY_TABLE = "adle_family_methods"
TEMPLATE_TABLE = "adle_activity_templates"
IMPORT_BATCH_TABLE = "canonical_teaching_dictionary_import_batches"


# ---------------------------------------------------------------------------
# Pure parsing + validation (regression-covered)
# ---------------------------------------------------------------------------

def split_sequence(raw: str, separator: str) -> list[str]:
    return [part.strip() for part in (raw or "").split(separator) if part.strip()]


def parse_workbook(workbook_path: Path) -> dict[str, Any]:
    from openpyxl import load_workbook

    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    result: dict[str, Any] = {"families": [], "templates": [], "errors": []}

    for sheet_name, headers in ((FAMILY_SHEET, FAMILY_HEADERS), (TEMPLATE_SHEET, TEMPLATE_HEADERS)):
        if sheet_name not in workbook.sheetnames:
            result["errors"].append(f"missing sheet: {sheet_name}")
            return result
        rows = list(workbook[sheet_name].iter_rows(values_only=True))
        actual_headers = tuple(str(cell).strip() if cell is not None else "" for cell in rows[0])
        if actual_headers != headers:
            result["errors"].append(
                f"{sheet_name}: header mismatch — expected {headers}, got {actual_headers}"
            )
            return result
        body = [row for row in rows[1:] if any(cell is not None and str(cell).strip() for cell in row)]
        if sheet_name == FAMILY_SHEET:
            for index, row in enumerate(body, start=2):
                values = {header: str(cell).strip() if cell is not None else "" for header, cell in zip(headers, row)}
                result["families"].append({
                    "source_row_number": index,
                    "family_key": values["family_key"],
                    "family_name": values["family_name"],
                    "core_pedagogy": values["core_pedagogy"],
                    "first_exposure_sequence": split_sequence(values["first_exposure_sequence"], ";"),
                    "guided_question_sequence": split_sequence(values["guided_question_sequence"], "->"),
                    "review_sort_dimension": values["review_sort_dimension"],
                    "production_task": values["production_task"],
                    "notes": values["notes"] or None,
                })
        else:
            for index, row in enumerate(body, start=2):
                values = {header: str(cell).strip() if cell is not None else "" for header, cell in zip(headers, row)}
                result["templates"].append({
                    "source_row_number": index,
                    "template_key": values["template_key"],
                    "phase": values["phase"],
                    "purpose": values["purpose"],
                    "child_response": values["child_response"],
                    "required_inputs": split_sequence(values["required_inputs"], ","),
                    "child_facing_copy": values["child_facing_copy"],
                })
    return result


def validate_registry(parsed: dict[str, Any]) -> list[str]:
    errors: list[str] = list(parsed.get("errors", []))
    families = parsed["families"]
    templates = parsed["templates"]

    if len(families) != EXPECTED_FAMILY_COUNT:
        errors.append(f"expected {EXPECTED_FAMILY_COUNT} family rows, found {len(families)}")
    if len(templates) != EXPECTED_TEMPLATE_COUNT:
        errors.append(f"expected {EXPECTED_TEMPLATE_COUNT} template rows, found {len(templates)}")

    family_keys = [family["family_key"] for family in families]
    if len(set(family_keys)) != len(family_keys):
        errors.append("duplicate family_key values on the Family Methods sheet")
    unexpected = sorted(set(family_keys) - EXPECTED_FAMILY_KEYS)
    missing = sorted(EXPECTED_FAMILY_KEYS - set(family_keys))
    if unexpected:
        errors.append(f"family keys not in the taxonomy skill_family_key set: {unexpected}")
    if missing:
        errors.append(f"taxonomy families missing from the sheet: {missing}")

    template_keys = [template["template_key"] for template in templates]
    if len(set(template_keys)) != len(template_keys):
        errors.append("duplicate template_key values on the Activity Templates sheet")
    template_key_set = set(template_keys)

    metadata_missing = sorted(template_key_set - set(TEMPLATE_RUNTIME_METADATA))
    metadata_extra = sorted(set(TEMPLATE_RUNTIME_METADATA) - template_key_set)
    if metadata_missing:
        errors.append(f"templates without declared runtime metadata: {metadata_missing}")
    if metadata_extra:
        errors.append(f"runtime metadata for unknown templates: {metadata_extra}")

    for template in templates:
        if template["phase"] not in KNOWN_PHASES:
            errors.append(
                f"unknown phase {template['phase']!r} on template {template['template_key']}"
            )
        for field in ("template_key", "phase", "purpose", "child_response", "child_facing_copy"):
            if not template[field]:
                errors.append(f"empty {field} on Activity Templates row {template['source_row_number']}")
        if not template["required_inputs"]:
            errors.append(f"empty required_inputs on template {template['template_key']}")

    for family in families:
        for field in ("family_key", "family_name", "core_pedagogy", "review_sort_dimension", "production_task"):
            if not family[field]:
                errors.append(f"empty {field} on Family Methods row {family['source_row_number']}")
        if not family["first_exposure_sequence"]:
            errors.append(f"empty first_exposure_sequence for family {family['family_key']}")
        if not family["guided_question_sequence"]:
            errors.append(f"empty guided_question_sequence for family {family['family_key']}")
        for key in family["guided_question_sequence"]:
            if key not in template_key_set and key not in SEQUENCE_META_KEYS:
                errors.append(
                    f"family {family['family_key']} guided sequence references unknown key {key!r}"
                )
    return errors


def build_report(parsed: dict[str, Any], errors: list[str], mode: str, workbook_path: Path) -> dict[str, Any]:
    return {
        "importer_version": IMPORTER_VERSION,
        "content_version": CONTENT_VERSION,
        "mode": mode,
        "input": str(workbook_path),
        "family_count": len(parsed["families"]),
        "template_count": len(parsed["templates"]),
        "family_keys": sorted(f["family_key"] for f in parsed["families"]),
        "sequence_meta_keys_allowed": sorted(SEQUENCE_META_KEYS),
        "validation_errors": errors,
        "batch_valid": not errors,
    }


# ---------------------------------------------------------------------------
# Guarded local-DB apply (guards match scripts/adle-band-teaching-dictionary.py)
# ---------------------------------------------------------------------------

def require_local_db_url(db_url: str) -> None:
    parsed = urlparse(db_url)
    if parsed.scheme not in {"postgres", "postgresql"}:
        raise ValueError("Local DB URL must use postgres:// or postgresql://.")
    if parsed.hostname not in {"127.0.0.1", "localhost"}:
        raise ValueError("Refusing non-local database host. Expected localhost or 127.0.0.1.")
    if parsed.port != 54322:
        raise ValueError("Refusing non-local Supabase port. Expected local Postgres port 54322.")
    if (parsed.path or "").lstrip("/") != "postgres":
        raise ValueError("Refusing non-local Supabase database. Expected database name postgres.")


def quote_sql_literal(value: str) -> str:
    return "'" + value.replace("'", "''") + "'"


def sql_text_array(values: list[str]) -> str:
    if not values:
        return "'{}'::text[]"
    return "array[" + ", ".join(quote_sql_literal(v) for v in values) + "]"


def psql_command_base(db_url: str, psql_command: str, psql_mode: str, docker_container: str | None) -> list[str]:
    if psql_mode == "host":
        return [psql_command, db_url, "--no-psqlrc", "--quiet", "--tuples-only", "--no-align", "-v", "ON_ERROR_STOP=1"]
    if psql_mode == "docker":
        if not docker_container:
            raise ValueError("--docker-container is required when --psql-mode docker is used.")
        return [
            "docker", "exec", "-i", docker_container,
            "psql", "-U", "postgres", "-d", "postgres",
            "--no-psqlrc", "--quiet", "--tuples-only", "--no-align", "-v", "ON_ERROR_STOP=1",
        ]
    raise ValueError(f"Unsupported psql mode: {psql_mode!r}.")


def run_psql_json(base: list[str], sql: str) -> Any:
    wrapped = (
        "select coalesce(jsonb_agg(row_to_json(result_rows)), '[]'::jsonb) "
        f"from ({sql}) result_rows"
    )
    result = subprocess.run([*base, "-c", wrapped], capture_output=True, text=True)
    if result.returncode != 0:
        detail = (result.stderr or result.stdout or "").strip()
        raise ValueError(detail or f"psql exited with status {result.returncode}")
    return json.loads(result.stdout.strip() or "[]")


def run_psql_script_text(base: list[str], sql: str) -> str:
    result = subprocess.run(base, input=sql, capture_output=True, text=True)
    if result.returncode != 0:
        detail = (result.stderr or result.stdout or "").strip()
        raise ValueError(detail or f"psql exited with status {result.returncode}")
    return result.stdout.strip()


def db_preflight(base: list[str]) -> None:
    versions = run_psql_json(
        base,
        "select version from supabase_migrations.schema_migrations where version = any (array["
        + ", ".join(quote_sql_literal(v) for v in EXPECTED_MIGRATION_VERSIONS)
        + "])",
    )
    present = {row["version"] for row in versions}
    missing = sorted(set(EXPECTED_MIGRATION_VERSIONS) - present)
    if missing:
        raise ValueError(f"Migration ledger is missing required versions: {missing}.")
    catalog_families = run_psql_json(
        base,
        "select distinct skill_family_key from public.micro_skill_catalog where is_active = true",
    )
    live_keys = {row["skill_family_key"] for row in catalog_families}
    if not EXPECTED_FAMILY_KEYS <= live_keys:
        raise ValueError(
            f"micro_skill_catalog is missing expected family keys: {sorted(EXPECTED_FAMILY_KEYS - live_keys)}."
        )


def apply_transaction_sql(parsed: dict[str, Any]) -> str:
    family_values = []
    for family in sorted(parsed["families"], key=lambda f: f["family_key"]):
        family_values.append(
            "("
            + ", ".join([
                quote_sql_literal(family["family_key"]),
                quote_sql_literal(family["family_name"]),
                quote_sql_literal(family["core_pedagogy"]),
                sql_text_array(family["first_exposure_sequence"]),
                sql_text_array(family["guided_question_sequence"]),
                quote_sql_literal(family["review_sort_dimension"]),
                quote_sql_literal(family["production_task"]),
                quote_sql_literal(family["notes"]) if family["notes"] else "null",
                quote_sql_literal(CONTENT_VERSION),
                "(select id from _adle_registry_batch)",
            ])
            + ")"
        )
    template_values = []
    for template in sorted(parsed["templates"], key=lambda t: t["template_key"]):
        min_words, sentence, contrast, evidence_kind = TEMPLATE_RUNTIME_METADATA[template["template_key"]]
        template_values.append(
            "("
            + ", ".join([
                quote_sql_literal(template["template_key"]),
                quote_sql_literal(template["phase"]),
                quote_sql_literal(template["purpose"]),
                quote_sql_literal(template["child_response"]),
                sql_text_array(template["required_inputs"]),
                quote_sql_literal(template["child_facing_copy"]),
                str(min_words),
                "true" if sentence else "false",
                "true" if contrast else "false",
                quote_sql_literal(evidence_kind),
                quote_sql_literal(CONTENT_VERSION),
                "(select id from _adle_registry_batch)",
            ])
            + ")"
        )

    summary_json = quote_sql_literal(json.dumps({
        "importer_version": IMPORTER_VERSION,
        "content_version": CONTENT_VERSION,
        "family_count": len(family_values),
        "template_count": len(template_values),
    }, sort_keys=True, separators=(",", ":")))

    return f"""
begin;

select pg_advisory_xact_lock(hashtext({quote_sql_literal(ADVISORY_LOCK_NAME)}));

create temporary table _adle_registry_batch (id uuid not null) on commit preserve rows;

with inserted as (
  insert into public.{IMPORT_BATCH_TABLE} (
    source_folder_path, validator_version, validation_summary, row_counts,
    readiness_summary, import_mode, batch_status, source_metadata,
    imported_by, imported_at
  )
  values (
    {quote_sql_literal("adle-composer-registry:" + CONTENT_VERSION)},
    {quote_sql_literal(IMPORTER_VERSION)},
    {summary_json}::jsonb,
    {quote_sql_literal(json.dumps({FAMILY_TABLE: len(family_values), TEMPLATE_TABLE: len(template_values)}, sort_keys=True))}::jsonb,
    '{{}}'::jsonb,
    'local_dev_import',
    'applied',
    {quote_sql_literal(json.dumps({"run_kind": "adle_composer_registry_import"}, sort_keys=True))}::jsonb,
    'adle_composer_registry_importer',
    timezone('utc', now())
  )
  returning id
)
insert into _adle_registry_batch(id) select id from inserted;

-- Re-import supersedes prior active registry rows and inserts fresh ones.
update public.{FAMILY_TABLE}
   set row_status = 'superseded', updated_at = timezone('utc', now())
 where row_status = 'active';

insert into public.{FAMILY_TABLE} (
  family_key, family_name, core_pedagogy, first_exposure_sequence,
  guided_question_sequence, review_sort_dimension, production_task, notes,
  content_version, import_batch_id
)
values
{",".join(family_values)};

update public.{TEMPLATE_TABLE}
   set row_status = 'superseded', updated_at = timezone('utc', now())
 where row_status = 'active';

insert into public.{TEMPLATE_TABLE} (
  template_key, phase, purpose, child_response, required_inputs,
  child_facing_copy, min_words_required, requires_sentence_context,
  requires_contrast_words, evidence_kind, content_version, import_batch_id
)
values
{",".join(template_values)};

do $$
declare
  family_count integer;
  template_count integer;
begin
  select count(*) into family_count from public.{FAMILY_TABLE}
   where import_batch_id = (select id from _adle_registry_batch);
  if family_count <> {len(family_values)} then
    raise exception 'adle family-method row-count verification failed';
  end if;
  select count(*) into template_count from public.{TEMPLATE_TABLE}
   where import_batch_id = (select id from _adle_registry_batch);
  if template_count <> {len(template_values)} then
    raise exception 'adle activity-template row-count verification failed';
  end if;
end $$;

commit;

select jsonb_build_object(
  'actual_import_run', true,
  'registry_batch_id', (select id::text from _adle_registry_batch),
  'family_rows_inserted', {len(family_values)},
  'template_rows_inserted', {len(template_values)},
  'status', 'local_registry_import_committed'
)::text;
""".strip()


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main() -> int:
    parser = argparse.ArgumentParser(
        description="Import the ADLE composer registry sheets (dry-run by default).",
    )
    parser.add_argument("--workbook", default=str(DEFAULT_WORKBOOK), help="Content workbook path.")
    parser.add_argument("--report", help="JSON batch report output path.")
    parser.add_argument("--local-db-url", help="Local Supabase Postgres URL (guarded; required for --apply).")
    parser.add_argument("--apply", action="store_true", help="Write registry rows (local DB mode only).")
    parser.add_argument(
        "--confirm-local-dev-registry",
        help=f"Required with --apply. Must equal {LOCAL_CONFIRMATION_TOKEN!r}.",
    )
    parser.add_argument("--psql-command", default="psql", help="psql executable for host mode.")
    parser.add_argument("--psql-mode", choices=("host", "docker"), default="host")
    parser.add_argument("--docker-container", help="Local Supabase DB container for docker psql mode.")
    args = parser.parse_args()

    workbook_path = Path(args.workbook).expanduser().resolve()
    if not workbook_path.is_file():
        print(f"Workbook not found: {workbook_path}", file=sys.stderr)
        return 2

    parsed = parse_workbook(workbook_path)
    errors = validate_registry(parsed)
    report = build_report(parsed, errors, "dry_run", workbook_path)

    if args.apply and not errors:
        if not args.local_db_url:
            print("Refusing --apply without --local-db-url.", file=sys.stderr)
            return 2
        if args.confirm_local_dev_registry != LOCAL_CONFIRMATION_TOKEN:
            print(
                f"Refusing local registry apply without --confirm-local-dev-registry {LOCAL_CONFIRMATION_TOKEN!r}.",
                file=sys.stderr,
            )
            return 2
        require_local_db_url(args.local_db_url)
        base = psql_command_base(args.local_db_url, args.psql_command, args.psql_mode, args.docker_container)
        db_preflight(base)
        output = run_psql_script_text(base, apply_transaction_sql(parsed))
        apply_result = json.loads(output.splitlines()[-1] if output else "{}")
        report["mode"] = "local_db_apply"
        report["apply_result"] = apply_result

    if args.report:
        report_path = Path(args.report)
        report_path.parent.mkdir(parents=True, exist_ok=True)
        report_path.write_text(json.dumps(report, indent=2, sort_keys=True) + "\n", encoding="utf-8")

    print(f"ADLE composer registry import ({report['mode']}) — {CONTENT_VERSION}")
    print(f"Families: {report['family_count']}  Templates: {report['template_count']}")
    if errors:
        print("VALIDATION ERRORS (batch fails closed):", file=sys.stderr)
        for error in errors:
            print(f"  {error}", file=sys.stderr)
        return 1
    if args.apply and report["mode"] != "local_db_apply":
        return 1
    return 0


if __name__ == "__main__":
    sys.exit(main())
